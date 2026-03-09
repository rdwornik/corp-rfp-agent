"""Detect green cells and pair them with questions.

Replicates the exact detection logic from src/rfp_excel_agent.py:
- Green = exact FF00FF00 (8-char ARGB) or 00FF00 (6-char RGB)
- Checks fgColor first, then bgColor as fallback
- Question column: header-based detection (requirement > question > description)
- Answer column: header-based detection (vendor response > BY response > ...)
- Scans any cell in a row for green, uses first green cell found
"""

import logging
from pathlib import Path
from typing import Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from corp_rfp_agent.agents.excel.models import GreenCell

logger = logging.getLogger(__name__)

# Green color constants -- must match src/rfp_excel_agent.py exactly
GREEN_COLOR = "FF00FF00"
GREEN_COLOR_SHORT = "00FF00"


def is_green_cell(cell) -> bool:
    """Check if a cell has exact green fill color (FF00FF00 ONLY).

    Replicates src/rfp_excel_agent.is_green_cell exactly.
    """
    if not cell or not hasattr(cell, "fill"):
        return False

    fill = cell.fill
    if fill is None:
        return False

    # Check fgColor
    if hasattr(fill, "fgColor") and fill.fgColor:
        fg = fill.fgColor
        if hasattr(fg, "rgb") and fg.rgb:
            rgb = str(fg.rgb).upper()
            if len(rgb) == 8:
                return rgb == GREEN_COLOR
            elif len(rgb) == 6:
                return rgb == GREEN_COLOR_SHORT

    # Check bgColor as fallback
    if hasattr(fill, "bgColor") and fill.bgColor:
        bg = fill.bgColor
        if hasattr(bg, "rgb") and bg.rgb:
            rgb = str(bg.rgb).upper()
            if len(rgb) == 8:
                return rgb == GREEN_COLOR
            elif len(rgb) == 6:
                return rgb == GREEN_COLOR_SHORT

    return False


def find_header_row(sheet) -> int:
    """Find the header row (first row with >= 3 non-empty text cells).

    Replicates src/rfp_excel_agent.find_header_row exactly.
    """
    for row_idx in range(1, min(20, sheet.max_row + 1)):
        row_values = []
        for col_idx in range(1, min(50, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value and str(cell.value).strip():
                row_values.append(str(cell.value).strip())
        if len(row_values) >= 3:
            return row_idx
    return 1


def detect_question_column(sheet, header_row: int) -> Tuple[Optional[int], Optional[str]]:
    """Detect question column using priority-based header matching.

    Priority:
    1. 'requirement', 'question', 'description'
    2. 'customer question', 'rfp question', 'functional requirement'
    3. First column with avg text length > 20 chars

    Replicates src/rfp_excel_agent.detect_question_column exactly.
    """
    priority1 = ["requirement", "question", "description"]
    priority2 = ["customer question", "rfp question", "functional requirement"]

    headers = {}
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        if cell.value:
            header_text = str(cell.value).strip()
            header_normalized = header_text.lower().replace("_", " ")
            headers[col_idx] = (header_text, header_normalized)

    for col_idx, (original, normalized) in headers.items():
        for pattern in priority1:
            if pattern in normalized:
                return col_idx, original

    for col_idx, (original, normalized) in headers.items():
        for pattern in priority2:
            if pattern in normalized:
                return col_idx, original

    # Fallback: first column with substantial text
    for col_idx in range(1, sheet.max_column + 1):
        total_len = 0
        count = 0
        for row_idx in range(header_row + 1, min(header_row + 20, sheet.max_row + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value and str(cell.value).strip():
                total_len += len(str(cell.value).strip())
                count += 1
        if count > 0 and (total_len / count) > 20:
            header_cell = sheet.cell(row=header_row, column=col_idx)
            header_name = str(header_cell.value).strip() if header_cell.value else f"Column {col_idx}"
            return col_idx, header_name

    return None, None


def detect_answer_column(
    sheet, header_row: int, question_col: int
) -> Tuple[Optional[int], Optional[str]]:
    """Detect answer column using priority-based header matching.

    Priority:
    1. 'vendor comment/answer/response'
    2. 'by comment/answer/response', 'blue yonder'
    3. 'supplier comment/answer/response'
    4. 'comment', 'answer', 'response'
    5. First empty column right of question column

    Replicates src/rfp_excel_agent.detect_answer_column exactly.
    """
    priority1 = ["vendor comment", "vendor answer", "vendor response"]
    priority2 = ["by comment", "by answer", "by response", "blue yonder"]
    priority3 = ["supplier comment", "supplier answer", "supplier response"]
    priority4 = ["comment", "answer", "response"]

    headers = {}
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        if cell.value:
            header_text = str(cell.value).strip()
            header_normalized = header_text.lower().replace("_", " ")
            headers[col_idx] = (header_text, header_normalized)

    for priority in [priority1, priority2, priority3, priority4]:
        for col_idx, (original, normalized) in headers.items():
            for pattern in priority:
                if pattern in normalized:
                    return col_idx, original

    # Fallback: first empty column right of question col
    for col_idx in range(question_col + 1, sheet.max_column + 2):
        is_empty = True
        for row_idx in range(header_row + 1, min(header_row + 10, sheet.max_row + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value and str(cell.value).strip():
                is_empty = False
                break
        if is_empty:
            header_cell = sheet.cell(row=header_row, column=col_idx)
            if header_cell.value:
                return col_idx, str(header_cell.value).strip()
            else:
                return col_idx, f"Column_{get_column_letter(col_idx)}"

    return None, None


class CellDetector:
    """Detects green cells and identifies question/answer column pairs."""

    def __init__(self, workbook_path: Path):
        self._path = workbook_path
        self._wb = load_workbook(str(workbook_path), keep_links=True)

    def scan_green_cells(
        self,
        sheet_name: Optional[str] = None,
    ) -> list[GreenCell]:
        """Find all green cells that need answers.

        Replicates src/rfp_excel_agent.scan_green_cells exactly:
        - Iterates all sheets (or specified sheet)
        - Per sheet: find header row, detect question/answer columns
        - Per row after header: check all cells for green, use first match
        - Extract question text from question column
        - Skip rows with no question text
        """
        green_cells = []

        sheets = (
            [self._wb[sheet_name]]
            if sheet_name
            else self._wb.worksheets
        )

        for ws in sheets:
            header_row = find_header_row(ws)
            question_col, question_col_name = detect_question_column(ws, header_row)
            answer_col, answer_col_name = detect_answer_column(
                ws, header_row, question_col or 1
            )

            if not question_col:
                continue

            for row_idx in range(header_row + 1, ws.max_row + 1):
                row_has_green = False
                green_col = None

                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if is_green_cell(cell):
                        row_has_green = True
                        green_col = col_idx
                        break

                if row_has_green:
                    question_cell = ws.cell(row=row_idx, column=question_col)
                    question_text = (
                        str(question_cell.value).strip() if question_cell.value else ""
                    )

                    if question_text:
                        green_cells.append(
                            GreenCell(
                                sheet_name=ws.title,
                                row=row_idx,
                                answer_col_idx=answer_col or green_col,
                                question_col_idx=question_col,
                                question_text=question_text,
                                header_row=header_row,
                                green_cell_col_idx=green_col,
                            )
                        )

        logger.info("Found %d green cells in %s", len(green_cells), self._path.name)
        return green_cells

    @property
    def workbook(self):
        return self._wb

    def close(self):
        self._wb.close()
