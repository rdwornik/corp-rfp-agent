"""Review Pack Generator -- creates Excel with questions + RAG answers.

Generates an Excel workbook with realistic RFP questions, answers them
via ChromaDB RAG, and formats for human review.

Usage:
  python src/generate_review_pack.py --family wms
  python src/generate_review_pack.py --family wms --count 40
  python src/generate_review_pack.py --all
  python src/generate_review_pack.py --family wms --output-dir data/kb/review_packs/
"""

import argparse
import hashlib
import json
import os
import random
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import yaml

try:
    from dotenv import load_dotenv
    load_dotenv(override=False)
except ImportError:
    pass

PROJECT_ROOT = Path(__file__).resolve().parents[1]
KB_DIR = PROJECT_ROOT / "data" / "kb"
VERIFIED_DIR = KB_DIR / "verified"
DRAFTS_DIR = KB_DIR / "drafts"
PROFILES_DIR = PROJECT_ROOT / "config" / "product_profiles" / "_effective"
DEFAULT_OUTPUT_DIR = KB_DIR / "review_packs"

GENERATOR_VERSION = "1.0"

# BY brand colors
BY_MIDNIGHT_BLUE = "000E4E"
BY_HORIZON_BLUE = "00B7F1"
LIGHT_GREY = "F2F2F2"
YELLOW_BG = "FFFF99"
RED_FONT = "CC0000"

# Topics for question generation
QUESTION_TOPICS = [
    "deployment", "integration", "security", "data_management",
    "scalability", "architecture", "ui", "compliance",
    "disaster_recovery", "monitoring",
]

QUESTION_GEN_PROMPT = """You are generating realistic RFP questions for a Blue Yonder product.

PRODUCT: {product_display_name}

KEY CAPABILITIES:
{key_facts}

APIS: {apis}
DEPLOYMENT: {deployment}
CLOUD-NATIVE: {cloud_native}

FORBIDDEN CLAIMS (system must NOT assert these):
{forbidden_claims}

COVERAGE GAPS (topics with NO KB entries — prioritize these):
{gaps}

Generate {count} realistic RFP questions. Requirements:
1. Mix technical, functional, consulting, and customer_executive categories
2. Distribute across topics: {topics}
3. Include {edge_count} edge-case questions that test forbidden claim boundaries
4. Include questions about coverage gap topics
5. Vary difficulty: simple factual, comparison, scenario-based

Return JSON array:
[{{"question": "...", "topic": "deployment", "category": "technical", "source": "profile|forbidden|gap|existing"}}]"""


# ---------------------------------------------------------------------------
# Profile + KB loading
# ---------------------------------------------------------------------------

def load_profile(family: str, profiles_dir: Path = PROFILES_DIR) -> dict:
    """Load effective product profile."""
    path = profiles_dir / f"{family}.yaml"
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def load_existing_questions(family: str,
                            verified_dir: Path = VERIFIED_DIR) -> list[dict]:
    """Load sample of existing verified questions for quality checking."""
    questions = []
    family_dir = verified_dir / family
    if not family_dir.exists():
        return questions
    for json_file in sorted(family_dir.glob("*.json")):
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                entry = json.load(f)
            questions.append({
                "question": entry.get("question", ""),
                "topic": _infer_topic(entry),
                "category": entry.get("category", "technical"),
                "source": "existing",
            })
        except (json.JSONDecodeError, OSError):
            continue
    return questions


def _infer_topic(entry: dict) -> str:
    """Infer topic from entry text using CAPABILITY_TOPICS keywords."""
    from kb_eval import CAPABILITY_TOPICS
    text = (entry.get("question", "") + " " + entry.get("answer", "")).lower()
    for topic, keywords in CAPABILITY_TOPICS.items():
        if any(kw in text for kw in keywords):
            return topic
    return "general"


def get_coverage_gaps(family: str, profiles_dir: Path = PROFILES_DIR,
                      verified_dir: Path = VERIFIED_DIR,
                      drafts_dir: Path = DRAFTS_DIR) -> list[str]:
    """Get coverage gaps for a family from kb_eval."""
    from kb_eval import load_all_entries, load_all_profiles, check_coverage
    entries = load_all_entries(family_filter=family,
                              verified_dir=verified_dir, drafts_dir=drafts_dir)
    profiles = load_all_profiles(profiles_dir)
    coverage = check_coverage(entries, profiles, active_only=False)
    family_cov = coverage.get(family, {})
    return family_cov.get("gaps", [])


# ---------------------------------------------------------------------------
# Question generation
# ---------------------------------------------------------------------------

def generate_questions(profile: dict, count: int = 35,
                       gaps: list[str] | None = None,
                       existing: list[dict] | None = None,
                       model: str = "gemini-flash") -> list[dict]:
    """Generate RFP questions from profile context.

    Mix: 40% profile, 30% forbidden-claim edge cases, 20% existing, 10% gaps.
    """
    from kb_extract_historical import call_llm

    forbidden = profile.get("forbidden_claims", [])
    key_facts = profile.get("key_facts", [])
    gaps = gaps or []

    # Calculate mix
    n_existing = min(int(count * 0.2), len(existing or []))
    n_llm = count - n_existing
    n_edge = max(2, int(n_llm * 0.3))

    prompt = QUESTION_GEN_PROMPT.format(
        product_display_name=profile.get("display_name",
                                         profile.get("product", "unknown")),
        key_facts="\n".join(f"- {kf}" for kf in key_facts[:15]) if key_facts else "(none)",
        apis=profile.get("apis", []),
        deployment=profile.get("deployment", []),
        cloud_native=profile.get("cloud_native", "unknown"),
        forbidden_claims="\n".join(f"- {fc}" for fc in forbidden[:15]) if forbidden else "(none)",
        gaps=", ".join(gaps) if gaps else "(none)",
        count=n_llm,
        topics=", ".join(QUESTION_TOPICS),
        edge_count=n_edge,
    )

    raw = call_llm(prompt, model=model)
    questions = _parse_questions(raw)

    # Add sampled existing questions
    if existing and n_existing > 0:
        sampled = random.sample(existing, min(n_existing, len(existing)))
        questions.extend(sampled)

    # Ensure fields
    for q in questions:
        q.setdefault("topic", "general")
        q.setdefault("category", "technical")
        q.setdefault("source", "profile")

    return questions[:count]


def _parse_questions(raw_text: str) -> list[dict]:
    """Parse LLM response into question list."""
    if not raw_text:
        return []
    text = re.sub(r'^```(?:json)?\s*', '', raw_text.strip(), flags=re.MULTILINE)
    text = re.sub(r'```\s*$', '', text.strip(), flags=re.MULTILINE).strip()

    for attempt in [text, raw_text.strip()]:
        try:
            data = json.loads(attempt)
            if isinstance(data, list):
                return [q for q in data if isinstance(q, dict) and "question" in q]
        except json.JSONDecodeError:
            pass
        m = re.search(r'(\[[\s\S]*\])', attempt)
        if m:
            try:
                data = json.loads(m.group(1))
                if isinstance(data, list):
                    return [q for q in data if isinstance(q, dict) and "question" in q]
            except json.JSONDecodeError:
                pass
    return []


# ---------------------------------------------------------------------------
# RAG answering
# ---------------------------------------------------------------------------

def _content_hash(text: str) -> str:
    """SHA256 hash of text (first 16 chars)."""
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:16]


def answer_question(question: str, family: str,
                    chroma_path: Optional[Path] = None,
                    top_k: int = 3) -> dict:
    """Answer a question via ChromaDB RAG.

    Returns dict with: answer, entry_ids, confidence, revision_hash.
    """
    if chroma_path is None:
        chroma_path = KB_DIR / "chroma_store"

    if not chroma_path.exists():
        return _empty_answer()

    try:
        import chromadb
        from chromadb.utils import embedding_functions

        ef = embedding_functions.SentenceTransformerEmbeddingFunction(
            model_name="BAAI/bge-large-en-v1.5"
        )
        client = chromadb.PersistentClient(path=str(chroma_path))
        collection = client.get_collection(
            name="rfp_knowledge_base",
            embedding_function=ef,
        )

        results = collection.query(
            query_texts=[question],
            n_results=top_k,
            where={"domain": family} if family else None,
        )

        if not results or not results["ids"] or not results["ids"][0]:
            return _empty_answer()

        ids = results["ids"][0]
        distances = results["distances"][0] if results.get("distances") else [999] * len(ids)
        metadatas = results["metadatas"][0] if results.get("metadatas") else [{}] * len(ids)

        best_dist = distances[0]
        best_meta = metadatas[0]
        confidence = max(0.0, 1.0 - best_dist / 2.0)

        answer = best_meta.get("canonical_answer", "")
        entry_ids = [m.get("kb_id", i) for i, m in zip(ids, metadatas)]

        if not answer or confidence < 0.2:
            return _empty_answer()

        return {
            "answer": answer,
            "entry_ids": entry_ids[:top_k],
            "confidence": round(confidence, 3),
            "revision_hash": _content_hash(answer),
        }

    except Exception:
        return _empty_answer()


def _empty_answer() -> dict:
    return {
        "answer": "",
        "entry_ids": [],
        "confidence": 0.0,
        "revision_hash": "",
    }


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def create_review_pack(family: str, questions: list[dict],
                       rag_results: list[dict],
                       profile: dict,
                       output_dir: Path = DEFAULT_OUTPUT_DIR,
                       model: str = "gemini-flash") -> Path:
    """Create the Excel review pack."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # --- Colors & styles ---
    header_fill = PatternFill(start_color=BY_MIDNIGHT_BLUE,
                              end_color=BY_MIDNIGHT_BLUE, fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill(start_color=LIGHT_GREY,
                           end_color=LIGHT_GREY, fill_type="solid")
    yellow_fill = PatternFill(start_color=YELLOW_BG,
                              end_color=YELLOW_BG, fill_type="solid")
    red_font = Font(name="Calibri", color=RED_FONT, italic=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    top_align = Alignment(vertical="top")

    # ===== Sheet 1: Questions & Answers =====
    ws = wb.active
    ws.title = "Questions & Answers"

    headers = ["#", "Question", "Topic", "Category", "Answer",
               "KB Entries Used", "Confidence", "Action",
               "Feedback", "Rob's Answer", "_revision_hash", "_entry_ids_json"]
    widths = [5, 50, 15, 15, 60, 20, 12, 14, 50, 60, 0, 0]

    # Header row
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        if width > 0:
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        else:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    # Freeze pane on row 1
    ws.freeze_panes = "A2"

    # Action dropdown
    dv = DataValidation(
        type="list",
        formula1='"APPROVE,UPDATE,REJECT,NEW,SKIP"',
        allow_blank=True,
    )
    dv.error = "Please select: APPROVE, UPDATE, REJECT, NEW, or SKIP"
    dv.errorTitle = "Invalid Action"
    ws.add_data_validation(dv)

    # Data rows
    for row_idx, (q, rag) in enumerate(zip(questions, rag_results), 2):
        answer = rag.get("answer", "")
        confidence = rag.get("confidence", 0.0)
        entry_ids = rag.get("entry_ids", [])
        revision_hash = rag.get("revision_hash", "")

        # Pre-fill action
        if not answer:
            action = "NEW"
        elif confidence >= 0.8:
            action = "APPROVE"
        else:
            action = "SKIP"

        row_data = [
            row_idx - 1,                                     # A: #
            q.get("question", ""),                            # B: Question
            q.get("topic", ""),                               # C: Topic
            q.get("category", ""),                            # D: Category
            answer or "No answer found in KB",                # E: Answer
            ", ".join(entry_ids),                             # F: KB Entries Used
            confidence,                                       # G: Confidence
            action,                                           # H: Action
            "",                                               # I: Feedback
            "",                                               # J: Rob's Answer
            revision_hash,                                    # K: _revision_hash
            json.dumps(entry_ids),                            # L: _entry_ids_json
        ]

        is_alt = (row_idx % 2 == 0)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_align if col_idx in (2, 5, 9, 10) else top_align

            # Alternating row color
            if is_alt and col_idx <= 10:
                cell.fill = alt_fill

            # Empty answer → red italic
            if col_idx == 5 and not answer:
                cell.font = red_font

            # Confidence coloring
            if col_idx == 7:
                if confidence >= 0.8:
                    cell.font = Font(color="006600")
                elif confidence >= 0.5:
                    cell.font = Font(color="996600")
                else:
                    cell.font = Font(color="CC0000")

            # Feedback + Rob's Answer columns → yellow background
            if col_idx in (9, 10):
                cell.fill = yellow_fill

        # Apply dropdown validation to action cell
        dv.add(ws.cell(row=row_idx, column=8))

    # ===== Sheet 2: General Feedback =====
    ws2 = wb.create_sheet("General Feedback")
    for col_idx, (header, width) in enumerate(
        [("#", 5), ("Feedback", 80), ("Scope", 20)], 1
    ):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    ws2.freeze_panes = "A2"

    # Placeholder row
    ws2.cell(row=2, column=1, value=1)
    placeholder = ws2.cell(row=2, column=2,
                           value="(Write general feedback here)")
    placeholder.font = Font(color="999999", italic=True)
    ws2.cell(row=2, column=3, value="product-wide")

    # ===== Sheet 3: _metadata (hidden) =====
    now_iso = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    today = datetime.now().strftime("%Y%m%d")
    pack_id = f"{family.upper()}_{today}_PACK{random.randint(1, 999):03d}"

    ws3 = wb.create_sheet("_metadata")
    metadata = {
        "pack_id": pack_id,
        "family": family,
        "generated_at": now_iso,
        "generator_version": GENERATOR_VERSION,
        "kb_snapshot_hash": _content_hash(now_iso + family),
        "question_count": len(questions),
        "profile_status": profile.get("_meta", {}).get("status", "unknown"),
        "model_used": model,
    }
    for row_idx, (key, value) in enumerate(metadata.items(), 1):
        ws3.cell(row=row_idx, column=1, value=key)
        ws3.cell(row=row_idx, column=2, value=str(value))
    ws3.sheet_state = "hidden"

    # Save
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{pack_id}.xlsx"
    output_path = output_dir / filename
    wb.save(str(output_path))

    return output_path


# ---------------------------------------------------------------------------
# Full pipeline
# ---------------------------------------------------------------------------

def generate_pack(family: str, count: int = 35,
                  model: str = "gemini-flash",
                  output_dir: Path = DEFAULT_OUTPUT_DIR,
                  profiles_dir: Path = PROFILES_DIR,
                  verified_dir: Path = VERIFIED_DIR,
                  drafts_dir: Path = DRAFTS_DIR,
                  chroma_path: Optional[Path] = None) -> Optional[Path]:
    """Full pipeline: generate questions, answer via RAG, create Excel."""
    profile = load_profile(family, profiles_dir)
    if not profile:
        print(f"[ERROR] No profile for '{family}'")
        return None

    print(f"[1/4] Loading context for {family}...")
    gaps = get_coverage_gaps(family, profiles_dir, verified_dir, drafts_dir)
    existing = load_existing_questions(family, verified_dir)

    print(f"  Coverage gaps: {gaps or '(none)'}")
    print(f"  Existing questions: {len(existing)}")

    print(f"[2/4] Generating {count} questions...")
    questions = generate_questions(profile, count, gaps, existing, model)
    print(f"  Generated {len(questions)} questions")

    if not questions:
        print("[WARN] No questions generated")
        return None

    print(f"[3/4] Answering via RAG...")
    rag_results = []
    for i, q in enumerate(questions):
        print(f"  Answering {i+1}/{len(questions)}...", end="\r")
        result = answer_question(q["question"], family, chroma_path)
        rag_results.append(result)
    print()

    answered = sum(1 for r in rag_results if r["answer"])
    print(f"  Answered: {answered}/{len(questions)}")

    print(f"[4/4] Creating Excel pack...")
    path = create_review_pack(family, questions, rag_results, profile,
                              output_dir, model)
    print(f"  Saved: {path}")

    return path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generate Review Pack -- Excel with questions + RAG answers",
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--family", help="Product family")
    group.add_argument("--all", action="store_true", dest="all_families",
                       help="Generate for all active families")
    parser.add_argument("--count", type=int, default=35,
                        help="Questions per family (default: 35)")
    parser.add_argument("--model", default="gemini-flash",
                        help="LLM model (default: gemini-flash)")
    parser.add_argument("--output-dir", type=str, default=None,
                        help="Output directory (default: data/kb/review_packs/)")
    args = parser.parse_args()

    output_dir = Path(args.output_dir) if args.output_dir else DEFAULT_OUTPUT_DIR

    if args.all_families:
        if not PROFILES_DIR.exists():
            print("[ERROR] No profiles directory")
            return 1
        for p in sorted(PROFILES_DIR.glob("*.yaml")):
            if p.name.startswith("."):
                continue
            with open(p, "r", encoding="utf-8") as f:
                prof = yaml.safe_load(f) or {}
            if prof.get("_meta", {}).get("status") == "active":
                generate_pack(p.stem, args.count, args.model, output_dir)
    else:
        generate_pack(args.family, args.count, args.model, output_dir)

    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
