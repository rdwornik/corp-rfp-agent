"""
kb_extract_historical.py  —  3-Stage RFP Extraction Pipeline

FUNDAMENTAL PRINCIPLE: Row = atomic unit.
  Question and answer always come from the SAME row, different columns.

Pipeline:
  Stage 1 — Structure Detection
    openpyxl prescan (zero LLM) -> Gemini Flash column detection -> Rob confirms

  Stage 2 — Row Extraction + Filtering
    a. Extract every data row programmatically (openpyxl)
    b. Programmatic pre-filter: skip empty answers, too-short text, yes/no numbers
    c. LLM content filter (batch 10): classify BY_PRODUCT_ANSWER vs CLIENT_DATA etc.
    d. Print stats: what was kept vs what was filtered

  Stage 3 — Anonymize + Classify + Interactive Review
    a. Anonymize: replace client name with [Customer]
    b. LLM classify (batch 10): category, subcategory, tags, solution_codes,
       question_generic (rewrite removing client-specific phrasing)
    c. Interactive terminal review: Y/N/E/A/Q/S per entry

  Stage 4 — Generate Canonical + Archive
    Write clean v2 entries (14 fields, exactly canonical_entry_v2.json spec)
    Archive: move file + structure.json + extracted.jsonl + update registry

Usage:
  python src/kb_extract_historical.py --family network
  python src/kb_extract_historical.py --family network --file "path/to.xlsx"
  python src/kb_extract_historical.py --family planning         # auto IMPROVE mode
  python src/kb_extract_historical.py --family wms --resume
"""

import os, sys, json, re, shutil, textwrap
from pathlib import Path
from datetime import date, datetime
from collections import Counter
from typing import Optional

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv
load_dotenv(PROJECT_ROOT / ".env", override=True)

from answer_selector import (
    NumpyEncoder,
    select_answer, print_improve_report, save_improve_report,
    LLM_BUDGET_PER_FILE,
)

# ============================================================
# PATHS + CONSTANTS
# ============================================================
HISTORICAL_DIR = PROJECT_ROOT / "data/kb/historical"
CANONICAL_DIR  = PROJECT_ROOT / "data/kb/canonical"
STAGING_DIR    = PROJECT_ROOT / "data/kb/staging"
ARCHIVE_DIR    = PROJECT_ROOT / "data/kb/archive"
SCHEMA_DIR     = PROJECT_ROOT / "data/kb/schema"
FAMILY_CONFIG  = SCHEMA_DIR / "family_config.json"
REGISTRY_PATH  = ARCHIVE_DIR / "archive_registry.json"

TODAY   = date.today().isoformat()
NOW_ISO = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
WRAP_W  = 72

CATEGORY_CODES = {
    "functional": "FUNC", "technical": "TECH", "security": "SEC",
    "deployment": "DEPL", "commercial": "COM",  "general":  "GEN",
}

# Answers that are clearly trivial/empty
TRIVIAL_ANSWERS = {
    "", "-", "--", "---", "n/a", "na", "tbd", "tbc", "x", "none",
    "/", "yes", "no", "ja", "nein", "oui", "non", "si",
    "y", "n", "true", "false",
}


# ============================================================
# SECTION 1 — LLM LAYER
# ============================================================
def _load_models() -> dict:
    from src.llm_router import MODELS
    return MODELS


def call_llm(prompt: str, model: str = "gemini-flash") -> str:
    """Plain LLM call — no RAG. Returns raw text response."""
    models = _load_models()
    cfg   = models.get(model, models["gemini-flash"])
    prov  = cfg["provider"]
    mname = cfg["name"]

    if prov == "google":
        from google import genai
        from google.genai import types
        client = genai.Client(api_key=os.environ.get("GEMINI_API_KEY"))
        r = client.models.generate_content(
            model=mname,
            contents=[{"role": "user", "parts": [{"text": prompt}]}],
            config=types.GenerateContentConfig(temperature=0.1, max_output_tokens=8192),
        )
        return r.text.strip() if r.text else ""

    elif prov == "anthropic":
        import anthropic
        c = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
        m = c.messages.create(model=mname, max_tokens=4096, temperature=0.1,
                              messages=[{"role": "user", "content": prompt}])
        return m.content[0].text.strip()

    else:  # OpenAI-compatible
        from openai import OpenAI
        urls = {
            "openai": None, "deepseek": "https://api.deepseek.com/v1",
            "moonshot": "https://api.moonshot.ai/v1", "together": "https://api.together.xyz/v1",
            "xai": "https://api.x.ai/v1", "perplexity": "https://api.perplexity.ai",
            "mistral": "https://api.mistral.ai/v1",
            "alibaba": "https://dashscope.aliyuncs.com/compatible-mode/v1",
            "zhipu": "https://api.z.ai/api/paas/v4/",
        }
        keys = {
            "openai": "OPENAI_API_KEY", "deepseek": "DEEPSEEK_API_KEY",
            "moonshot": "MOONSHOT_API_KEY", "together": "TOGETHER_API_KEY",
            "xai": "XAI_API_KEY", "perplexity": "PERPLEXITY_API_KEY",
            "mistral": "MISTRAL_API_KEY", "alibaba": "DASHSCOPE_API_KEY",
            "zhipu": "ZHIPU_API_KEY",
        }
        kw = {"api_key": os.environ.get(keys.get(prov, ""))}
        if urls.get(prov):
            kw["base_url"] = urls[prov]
        c = OpenAI(**kw)
        r = c.chat.completions.create(model=mname, temperature=0.1, max_tokens=4096,
                                      messages=[{"role": "user", "content": prompt}])
        return r.choices[0].message.content.strip()


def call_llm_json(prompt: str, model: str) -> object:
    """Call LLM expecting JSON back. Strips code fences and parses."""
    raw = call_llm(prompt, model)
    raw = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.MULTILINE)
    raw = re.sub(r"```\s*$", "", raw.strip(), flags=re.MULTILINE).strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        m = re.search(r"(\{[\s\S]*\}|\[[\s\S]*\])", raw)
        if m:
            return json.loads(m.group(1))
        raise ValueError(f"LLM returned non-JSON:\n{raw[:400]}")


# ============================================================
# SECTION 2 — FAMILY + CANONICAL + ARCHIVE HELPERS
# ============================================================
def load_family_config() -> dict:
    with open(FAMILY_CONFIG, "r", encoding="utf-8") as f:
        return json.load(f)["families"]


def get_family(key: str) -> dict:
    cfg = load_family_config()
    if key not in cfg:
        raise ValueError(f"Unknown family '{key}'. Valid: {list(cfg.keys())}")
    return cfg[key]


def load_canonical(canon_file: str) -> list:
    p = CANONICAL_DIR / canon_file
    if not p.exists():
        return []
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)


def save_canonical(canon_file: str, entries: list) -> None:
    p = CANONICAL_DIR / canon_file
    with open(p, "w", encoding="utf-8") as f:
        json.dump(entries, f, indent=2, ensure_ascii=False, cls=NumpyEncoder)
    print(f"[OK] Wrote {len(entries)} entries -> {p.name}")


def next_seq(existing: list, prefix: str) -> int:
    """Return next sequence number after the highest existing one for this prefix."""
    pat = re.compile(rf"^{re.escape(prefix)}-\w+-(\d+)$")
    mx = 0
    for e in existing:
        m = pat.match(e.get("id", "") or e.get("kb_id", ""))
        if m:
            mx = max(mx, int(m.group(1)))
    return mx + 1


def load_registry() -> dict:
    if not REGISTRY_PATH.exists():
        return {"version": "1.0", "last_updated": "", "total_files": 0,
                "total_qa_extracted": 0, "entries": []}
    with open(REGISTRY_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_registry(reg: dict) -> None:
    reg["last_updated"] = NOW_ISO
    with open(REGISTRY_PATH, "w", encoding="utf-8") as f:
        json.dump(reg, f, indent=2, ensure_ascii=False, cls=NumpyEncoder)


def next_archive_id(reg: dict) -> str:
    nums = [int(re.search(r"\d+", e["archive_id"]).group())
            for e in reg.get("entries", []) if re.search(r"\d+", e.get("archive_id",""))]
    return f"ARC-{(max(nums) + 1 if nums else 1):04d}"


def make_archived_filename(metadata: dict, family_key: str, ext: str = ".xlsx") -> str:
    date_str = (metadata.get("date_estimated") or "unknown").replace(" ", "")
    client   = re.sub(r"[^\w\s]", "", metadata.get("client", "unknown")).strip()
    client   = re.sub(r"\s+", "_", client) or "unknown"
    rtype    = metadata.get("rfp_type", "response")
    return f"{date_str}_{client}_{family_key.upper()}_{rtype}{ext}"


# ============================================================
# SECTION 3 — METADATA AUTO-DETECTION
# ============================================================
def parse_filename_metadata(filename: str) -> dict:
    stem = Path(filename).stem
    name = re.sub(r"[_\-]+", " ", stem)

    yr_m  = re.search(r"\b(20\d{2})\b", name)
    year  = yr_m.group(1) if yr_m else None

    q_m = re.search(r"Q([1-4])", name, re.IGNORECASE)
    if q_m:
        quarter = f"Q{q_m.group(1)}"
    else:
        month_map = {"jan":"Q1","feb":"Q1","mar":"Q1","apr":"Q2","may":"Q2","jun":"Q2",
                     "jul":"Q3","aug":"Q3","sep":"Q3","oct":"Q4","nov":"Q4","dec":"Q4"}
        mo_m  = re.search(r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b",
                          name, re.IGNORECASE)
        quarter = month_map.get(mo_m.group(1).lower()) if mo_m else None

    date_est = f"{year}-{quarter}" if (year and quarter) else (year or "")

    nl = name.lower()
    if any(k in nl for k in ["response", "answer", "proposal"]):
        rfp_type = "response"
    elif any(k in nl for k in ["source", "request", "rfp", "rfi", "rfq", "rft"]):
        rfp_type = "source"
    else:
        rfp_type = "response"

    clean = re.sub(r"\b20\d{2}\b", "", name)
    clean = re.sub(r"\bQ[1-4]\b", "", clean, flags=re.IGNORECASE)
    for kw in ["rfp","rfi","rft","rfq","response","source","combined","wms","tms",
               "planning","logistics","scpo","catman","workforce","commerce","flexis",
               "network","doddle","aiml","blue yonder","by","question","answer",
               "proposal","request","tender","template","scm","inbound","tool"]:
        clean = re.sub(rf"\b{re.escape(kw)}\b", "", clean, flags=re.IGNORECASE)
    words  = [w for w in clean.split() if len(w) > 2]
    client = " ".join(words[:3]).strip()

    return {"client": client, "date_estimated": date_est, "rfp_type": rfp_type}


# ============================================================
# SECTION 4 — STAGE 1: STRUCTURE DETECTION
# ============================================================
def prescan_excel(filepath: Path) -> dict:
    """Read first 20 rows per sheet. Zero LLM cost."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {"filename": filepath.name, "sheets": []}
    for sname in wb.sheetnames:
        ws     = wb[sname]
        rows   = []
        merged = []
        for row in ws.iter_rows(max_row=20):
            row_data = {}
            for cell in row:
                if cell.value is not None:
                    row_data[cell.column_letter] = {
                        "value": str(cell.value)[:200],
                        "row":   cell.row,
                    }
            if row_data:
                rows.append(row_data)
        try:
            merged = [str(m) for m in ws.merged_cells.ranges]
        except Exception:
            merged = []
        result["sheets"].append({
            "name":        sname,
            "total_rows":  ws.max_row or 0,
            "total_cols":  ws.max_column or 0,
            "sample_rows": rows,
            "merged_cells": merged[:20],
        })
    wb.close()
    return result


def collect_metadata_interactive(filename: str, family_key: str) -> dict:
    guess = parse_filename_metadata(filename)
    print(f"\n{'='*60}")
    print(f" New file: {filename}")
    print(f" Family  : {family_key.upper()}")
    print(f"{'='*60}")

    def ask(label: str, default: str, options: str = "") -> str:
        opts = f"  ({options})" if options else ""
        val  = input(f"  {label} [{default}]{opts}: ").strip()
        return val if val else default

    _DATE_PATTERNS = [
        re.compile(r"^\d{4}-Q[1-4]$"),
        re.compile(r"^\d{4}-\d{2}$"),
        re.compile(r"^\d{4}-\d{2}-\d{2}$"),
    ]

    def ask_date(label: str, default: str) -> str:
        while True:
            val = input(f"  {label} [{default}]: ").strip()
            if not val:
                return default
            if any(p.match(val) for p in _DATE_PATTERNS):
                return val
            print("  [!] Use YYYY-QN (2024-Q3), YYYY-MM (2026-02), or YYYY-MM-DD (2026-02-19).")

    client   = ask("Client name", guess["client"] or "unknown")
    industry = ask("Industry", "retail",
                   "retail/cpg/manufacturing/3pl/auto/pharma/fmcg/grocery/fashion/other")
    date_est = ask_date("Date (YYYY-QN, YYYY-MM, or YYYY-MM-DD)",
                        guess["date_estimated"] or "unknown")
    region   = ask("Region", "EMEA", "EMEA/NA/APAC/LATAM")
    rtype    = ask("File type", guess["rfp_type"], "source/response/combined")
    notes    = ask("Notes (optional)", "")

    return {
        "client": client, "client_industry": industry, "date_estimated": date_est,
        "region": region, "rfp_type": rtype, "notes": notes,
    }


def analyze_structure_llm(prescan: dict, family_display: str, model: str) -> dict:
    """One LLM call to detect column layout across all sheets."""
    compact = {"filename": prescan["filename"], "sheets": []}
    for sh in prescan["sheets"]:
        rows_text = []
        for row in sh["sample_rows"][:15]:
            row_text = "  ".join(
                f"{col}:{info['value'][:80]}" for col, info in sorted(row.items())
            )
            rows_text.append(f"Row{list(row.values())[0]['row']}: {row_text}")
        compact["sheets"].append({
            "name":        sh["name"],
            "total_rows":  sh["total_rows"],
            "merged_cells": sh["merged_cells"][:5],
            "sample":      rows_text,
        })

    prompt = f"""Analyze this RFP Excel file for Blue Yonder {family_display}.

File contents (first 15 rows per sheet, columns labeled A/B/C...):
{json.dumps(compact, indent=2, cls=NumpyEncoder)}

This RFP may contain client questions, BY product answers, requirement IDs,
category headers, compliance indicators (Y/N/Partial), and comments.

IMPORTANT: Question and answer are ALWAYS in the same row, different columns.

Return ONLY a JSON object:
{{
  "relevant_sheets": [
    {{
      "sheet_name": "...",
      "purpose": "questions_and_answers" | "questions_only" | "metadata" | "skip",
      "data_start_row": <int>,
      "columns": {{
        "question_id":  "<col_letter or null>",
        "category":     "<col_letter or null>",
        "question":     "<col_letter>",
        "answer":       "<col_letter or null>",
        "compliance":   "<col_letter or null>",
        "comments":     "<col_letter or null>"
      }},
      "notes": "1 sentence"
    }}
  ],
  "file_type": "source_rfp" | "response" | "combined" | "unknown",
  "estimated_questions": <int>
}}

Rules:
- Include only sheets with actual RFP content; mark cover/TOC/scoring as "skip"
- "question" column is required for relevant sheets
- "answer" is null for source_rfp files (no BY responses yet)
- Return ONLY valid JSON, no other text"""

    return call_llm_json(prompt, model)


def confirm_structure(structure: dict, filename: str) -> bool:
    """Print detected structure, ask Rob to confirm."""
    print(f"\n--- Structure: {filename} ---")
    sheets = structure.get("relevant_sheets", [])
    if not sheets:
        print("  No relevant sheets detected.")
        return False

    for sh in sheets:
        if sh.get("purpose") == "skip":
            print(f"  [SKIP] {sh['sheet_name']}")
            continue
        cols = sh.get("columns", {})
        print(f"  Sheet '{sh['sheet_name']}' ({sh.get('purpose','?')}, "
              f"{sh.get('data_start_row','?')}+ rows):")
        print(f"    Q={cols.get('question','?')}  "
              f"A={cols.get('answer','-')}  "
              f"Cat={cols.get('category','-')}")
        if sh.get("notes"):
            print(f"    Note: {sh['notes']}")

    est = structure.get("estimated_questions", "?")
    print(f"\n  File type : {structure.get('file_type','unknown')}")
    print(f"  Est. pairs: {est}")

    ans = input("\n  Proceed? [Y/n]: ").strip().lower()
    return ans in ("", "y", "yes")


# ============================================================
# SECTION 5 — STAGE 2a: ROW EXTRACTION
# ============================================================
def _col_idx(letter: Optional[str]) -> int:
    """Convert column letter (A, B, AA…) to 0-based index."""
    if not letter:
        return -1
    letter = letter.strip().upper()
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def extract_all_rows(filepath: Path, structure: dict) -> list[dict]:
    """
    Extract every data row from relevant sheets.
    Returns raw rows — no filtering yet.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    all_rows = []
    row_num_global = 0

    for sh_map in structure.get("relevant_sheets", []):
        if sh_map.get("purpose") == "skip":
            continue
        sname = sh_map["sheet_name"]
        if sname not in wb.sheetnames:
            continue

        ws      = wb[sname]
        cols    = sh_map.get("columns", {})
        q_col   = cols.get("question")
        a_col   = cols.get("answer")
        cat_col = cols.get("category")
        start   = max(int(sh_map.get("data_start_row", 2)), 2)

        current_category = ""

        for row_idx, row_tuple in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx < start:
                continue

            def cell(letter):
                idx = _col_idx(letter)
                if idx < 0 or idx >= len(row_tuple):
                    return ""
                v = row_tuple[idx]
                return str(v).strip() if v is not None else ""

            q_text   = cell(q_col)
            a_text   = cell(a_col)
            cat_text = cell(cat_col)

            # Track running category from header rows (short text, no answer)
            if cat_text and not q_text:
                current_category = cat_text
                continue
            if q_text and len(q_text) < 50 and not a_text:
                current_category = q_text
                continue

            if not q_text:
                continue

            row_num_global += 1
            all_rows.append({
                "row_num":          row_num_global,
                "source_sheet":     sname,
                "source_row":       row_idx,
                "source_file":      filepath.name,
                "category_hint":    cat_text or current_category,
                "question":         q_text,
                "answer":           a_text,
            })

    wb.close()
    return all_rows


# ============================================================
# SECTION 6 — STAGE 2b: PROGRAMMATIC FILTER
# ============================================================
def programmatic_filter(rows: list[dict]) -> tuple[list[dict], dict]:
    """
    Fast heuristic pre-filter before LLM calls.
    Returns (candidates, stats).
    """
    stats      = Counter()
    candidates = []

    for row in rows:
        q = row["question"]
        a = row["answer"]

        if not q or len(q) < 10:
            stats["skip_short_question"] += 1
            continue

        if not a or a.lower() in TRIVIAL_ANSWERS:
            stats["skip_empty_answer"] += 1
            continue

        # Answer is a short number or single word (likely client data)
        a_clean = a.replace(",", "").replace(".", "").replace("%", "").replace(" ", "")
        if len(a) <= 10 and (a_clean.isdigit() or a.lower() in TRIVIAL_ANSWERS):
            stats["skip_client_data_short"] += 1
            continue

        candidates.append(row)

    stats["candidates"] = len(candidates)
    return candidates, stats


# ============================================================
# SECTION 7 — STAGE 2c: LLM CONTENT FILTER
# ============================================================
def llm_content_filter(candidates: list[dict], family_display: str,
                       model: str) -> tuple[list[dict], Counter]:
    """
    LLM decides: BY_PRODUCT_ANSWER vs CLIENT_DATA / CUSTOMER_SPECIFIC / INSTRUCTIONS.
    Batch size 10. Returns (kept_rows, classification_counts).
    """
    BATCH = 10
    kept   = []
    counts = Counter()

    for i in range(0, len(candidates), BATCH):
        batch      = candidates[i : i + BATCH]
        batch_data = [
            {
                "row_num":  r["row_num"],
                "question": r["question"][:250],
                "answer":   r["answer"][:200],
            }
            for r in batch
        ]

        prompt = f"""You are filtering rows from a Blue Yonder {family_display} RFP document.

Each row contains a question and an answer. Classify each one:

  BY_PRODUCT_ANSWER  — Blue Yonder describes its own product capability, feature,
                       architecture, or process. The answer is substantive (multiple
                       sentences describing what BY does). THIS IS WHAT WE WANT.

  CLIENT_DATA        — The client describes their own company: headcount, locations,
                       system landscape, volumes, current processes, budgets.
                       Answers are typically short (a number, a list, yes/no).

  CUSTOMER_SPECIFIC  — The question is so tied to this specific client that it cannot
                       be reused (e.g. "List your contracts with Retailer X in 2023").

  INSTRUCTIONS       — Instructions for filling in the RFP, notes, table headers,
                       formatting guidance.

  UNCLEAR            — Cannot determine; default to skip.

Return a JSON array — one object per row, same order as input:
[{{
  "row_num": <int>,
  "classification": "BY_PRODUCT_ANSWER" | "CLIENT_DATA" | "CUSTOMER_SPECIFIC" | "INSTRUCTIONS" | "UNCLEAR",
  "keep": true | false,
  "reason": "3-5 words"
}}]

RULE: keep=true ONLY for BY_PRODUCT_ANSWER.

Rows to classify:
{json.dumps(batch_data, indent=2, cls=NumpyEncoder)}

Return ONLY valid JSON array."""

        try:
            result = call_llm_json(prompt, model)
            if not isinstance(result, list):
                raise ValueError("Expected JSON array")

            keep_nums = {
                item["row_num"]
                for item in result
                if item.get("keep") is True
            }
            for item in result:
                cls = item.get("classification", "UNCLEAR").upper()
                counts[cls] += 1

            for row in batch:
                if row["row_num"] in keep_nums:
                    kept.append(row)

        except Exception as e:
            print(f"\n  [WARN] LLM content filter batch failed: {e}")
            # Conservative fallback: keep the batch
            kept.extend(batch)
            counts["UNCLEAR"] += len(batch)

        done = min(i + BATCH, len(candidates))
        print(f"  Content filter: {done}/{len(candidates)}...", end="\r")

    print()
    return kept, counts


def print_filter_stats(total_rows: int, prog_stats: Counter, llm_counts: Counter,
                       kept: int, sheet_name: str = "") -> None:
    label = f" ({sheet_name})" if sheet_name else ""
    print(f"\n  Row analysis{label} — {total_rows} total rows:")
    print(f"    [+] BY product answers (kept)  : {kept}")
    print(f"    [-] Empty / trivial answers    : {prog_stats.get('skip_empty_answer', 0)}")
    print(f"    [-] Short question / no text   : {prog_stats.get('skip_short_question', 0)}")
    print(f"    [-] Short number / client value: {prog_stats.get('skip_client_data_short', 0)}")
    if llm_counts:
        print(f"    [-] CLIENT_DATA (LLM)          : {llm_counts.get('CLIENT_DATA', 0)}")
        print(f"    [-] CUSTOMER_SPECIFIC (LLM)    : {llm_counts.get('CUSTOMER_SPECIFIC', 0)}")
        print(f"    [-] INSTRUCTIONS (LLM)         : {llm_counts.get('INSTRUCTIONS', 0)}")
        print(f"    [-] UNCLEAR / other (LLM)      : {llm_counts.get('UNCLEAR', 0)}")
    print(f"    -> {kept} rows to review")


# ============================================================
# SECTION 8 — STAGE 3a: ANONYMIZE
# ============================================================
def anonymize_rows(rows: list[dict], client_name: str) -> list[dict]:
    """Replace client name (and variations) with [Customer] in question + answer."""
    if not client_name or client_name.lower() in ("unknown", ""):
        return rows

    # Build patterns: exact name + common variations
    name_parts = client_name.strip().split()
    patterns   = [re.escape(client_name)]
    if len(name_parts) > 1:
        patterns.append(re.escape(name_parts[0]))  # first word only

    combined = re.compile("|".join(patterns), re.IGNORECASE)

    result = []
    for row in rows:
        row = dict(row)
        row["question"] = combined.sub("[Customer]", row["question"])
        row["answer"]   = combined.sub("[Customer]", row["answer"])
        result.append(row)
    return result


# ============================================================
# SECTION 9 — STAGE 3b: CLASSIFY
# ============================================================
def classify_rows_batch(rows: list[dict], family: dict, model: str) -> list[dict]:
    """
    LLM classifies each kept row.
    Adds: category, subcategory, tags, solution_codes, question_variants, question_generic.
    question_generic = question rewritten to remove client-specific phrasing.
    Batch size 10.
    """
    BATCH    = 10
    sol_list = ", ".join(family["solution_codes"])
    fname    = family["display_name"]
    result   = []

    for i in range(0, len(rows), BATCH):
        batch = rows[i : i + BATCH]
        items = ""
        for j, r in enumerate(batch):
            items += (
                f'\n[{j}]\n'
                f'Q: {r["question"][:300]}\n'
                f'A: {r["answer"][:200]}\n'
            )

        prompt = f"""Classify these Q&A pairs from a Blue Yonder {fname} RFP.

Available solution codes: {sol_list}

Return a JSON array — one object per pair, same order:
[{{
  "idx": <int>,
  "category": "functional|technical|security|deployment|commercial|general",
  "subcategory": "<snake_case topic, e.g. wave_picking, sso_auth, carrier_integration>",
  "tags": ["keyword1", "keyword2", "keyword3"],
  "solution_codes": [<applicable codes, or [] if all solutions in family>],
  "question_variants": ["alt phrasing 1", "alt phrasing 2"],
  "question_generic": "<question rewritten without client-specific details; keep as-is if already generic>"
}}]

Category guide:
  functional  — Business capabilities, workflows, features, UI, processes, reporting
  technical   — Architecture, APIs, integrations, data model, performance, platform
  security    — Authentication, encryption, compliance certs, access control, data residency
  deployment  — SaaS hosting, environments, upgrade cadence, SLAs, go-live, implementation
  commercial  — BY licensing model (NOT client volumes, revenue, or headcount)
  general     — Company overview, references (ONLY if nothing else fits)

Q&A pairs:
{items}

Return ONLY valid JSON array."""

        try:
            res = call_llm_json(prompt, model)
            if isinstance(res, list):
                for item in res:
                    idx = item.get("idx", -1)
                    if 0 <= idx < len(batch):
                        batch[idx].update({
                            "category":          item.get("category", "functional"),
                            "subcategory":       item.get("subcategory", ""),
                            "tags":              item.get("tags", []),
                            "solution_codes":    item.get("solution_codes", []),
                            "question_variants": item.get("question_variants", []),
                            "question_generic":  item.get("question_generic", ""),
                        })
        except Exception as e:
            print(f"\n  [WARN] Classify batch {i//BATCH+1} failed: {e}")
            for r in batch:
                r.setdefault("category", "functional")
                r.setdefault("subcategory", "")
                r.setdefault("tags", [])
                r.setdefault("solution_codes", [])
                r.setdefault("question_variants", [])
                r.setdefault("question_generic", "")

        result.extend(batch)
        print(f"  Classifying: {min(i+BATCH, len(rows))}/{len(rows)}...", end="\r")

    print()
    return result


def classify_rows_batch_async(rows: list[dict], family: dict, model: str) -> list[dict]:
    """Classify rows via Gemini Batch API (50% cheaper, no rate limits).

    Same logic as classify_rows_batch but submits all prompt-batches as one
    async batch job instead of sequential synchronous calls.
    """
    from batch_llm import BatchProcessor, parse_json_from_batch

    BATCH    = 10
    sol_list = ", ".join(family["solution_codes"])
    fname    = family["display_name"]

    processor = BatchProcessor(model=model)
    batches   = []

    for i in range(0, len(rows), BATCH):
        batch = rows[i : i + BATCH]
        items = ""
        for j, r in enumerate(batch):
            items += (
                f'\n[{j}]\n'
                f'Q: {r["question"][:300]}\n'
                f'A: {r["answer"][:200]}\n'
            )

        prompt = f"""Classify these Q&A pairs from a Blue Yonder {fname} RFP.

Available solution codes: {sol_list}

Return a JSON array -- one object per pair, same order:
[{{
  "idx": <int>,
  "category": "functional|technical|security|deployment|commercial|general",
  "subcategory": "<snake_case topic>",
  "tags": ["keyword1", "keyword2", "keyword3"],
  "solution_codes": [<applicable codes, or [] if all solutions in family>],
  "question_variants": ["alt phrasing 1", "alt phrasing 2"],
  "question_generic": "<question rewritten without client-specific details>"
}}]

Category guide:
  functional  -- Business capabilities, workflows, features, UI, processes, reporting
  technical   -- Architecture, APIs, integrations, data model, performance, platform
  security    -- Authentication, encryption, compliance certs, access control
  deployment  -- SaaS hosting, environments, upgrade cadence, SLAs, go-live
  commercial  -- BY licensing model (NOT client volumes, revenue)
  general     -- Company overview, references (ONLY if nothing else fits)

Q&A pairs:
{items}

Return ONLY valid JSON array."""

        processor.add(key=f"batch_{i}", prompt=prompt)
        batches.append((i, batch))

    print(f"  Submitting {processor.count} classification batches to Batch API...")
    result = processor.run(display_name="kb-classify", verbose=True)
    print(f"  Batch API: {result.succeeded} succeeded, {result.failed} failed")

    # Apply results
    all_rows = []
    for i, batch in batches:
        key = f"batch_{i}"
        if key in result.results:
            try:
                res = parse_json_from_batch(result.results[key])
                if isinstance(res, list):
                    for item in res:
                        idx = item.get("idx", -1)
                        if 0 <= idx < len(batch):
                            batch[idx].update({
                                "category":          item.get("category", "functional"),
                                "subcategory":       item.get("subcategory", ""),
                                "tags":              item.get("tags", []),
                                "solution_codes":    item.get("solution_codes", []),
                                "question_variants": item.get("question_variants", []),
                                "question_generic":  item.get("question_generic", ""),
                            })
            except (ValueError, Exception) as e:
                print(f"  [WARN] Parse failed for {key}: {e}")

        # Default any unclassified rows
        for r in batch:
            r.setdefault("category", "functional")
            r.setdefault("subcategory", "")
            r.setdefault("tags", [])
            r.setdefault("solution_codes", [])
            r.setdefault("question_variants", [])
            r.setdefault("question_generic", "")

        all_rows.extend(batch)

    return all_rows


# ============================================================
# SECTION 10 — STAGE 3c: INTERACTIVE REVIEW
# ============================================================
def _wrap(text: str, indent: int = 4) -> str:
    prefix = " " * indent
    return textwrap.fill(str(text), width=WRAP_W,
                         initial_indent=prefix, subsequent_indent=prefix)


def _print_review_card(row: dict, idx: int, total: int) -> None:
    print(f"\n{'='*60}")
    print(f"  Review | {row.get('source_file','?')} | {idx+1}/{total}")
    print(f"  Cat   : {row.get('category','?')} > {row.get('subcategory','?')}")
    print(f"{'='*60}")

    generic = row.get("question_generic", "").strip()
    original = row["question"].strip()

    print()
    if generic and generic != original:
        print("  Q (generic):")
        print(_wrap(generic))
        print("  Q (original):")
        print(_wrap(original))
    else:
        print("  Q:")
        print(_wrap(original))

    print()
    print("  A:")
    print(_wrap(row["answer"][:500]))

    print()
    tags = ", ".join(row.get("tags", [])) or "-"
    sols = ", ".join(row.get("solution_codes", [])) or "all"
    print(f"  Tags     : {tags}")
    print(f"  Solutions: {sols}")
    print()
    print("-" * 60)
    print("  [Y/Enter] Accept   [N] Skip   [E] Edit")
    print("  [A] Accept all     [S] Stats  [Q] Quit & save")
    print("-" * 60)


def _edit_row(row: dict) -> dict:
    print("\n  -- EDIT --")

    def ef(label: str, current: str) -> str:
        preview = current[:80] + ("..." if len(current) > 80 else "")
        print(f"  {label} [{preview}]")
        v = input("  New value (Enter to keep): ").strip()
        return v if v else current

    row = dict(row)
    generic  = row.get("question_generic", "") or row["question"]
    new_q    = ef("Question", generic)
    row["question_generic"] = new_q
    row["answer"]           = ef("Answer", row["answer"])

    tag_in = input(f"  Tags [{', '.join(row.get('tags',[]))}] (Enter to keep): ").strip()
    if tag_in:
        row["tags"] = [t.strip() for t in tag_in.split(",") if t.strip()]
    return row


def _print_stats(accepted: int, skipped: int, total: int) -> None:
    print(f"\n  Accepted: {accepted}  Skipped: {skipped}  Remaining: {total-accepted-skipped}")


def session_path(family_key: str) -> Path:
    STAGING_DIR.mkdir(exist_ok=True)
    return STAGING_DIR / f"{family_key}_session.json"


def save_session(family_key: str, filename: str, reviewed: list,
                 next_idx: int, metadata: dict, structure: dict) -> None:
    data = {
        "family": family_key, "file": filename,
        "metadata": metadata, "structure": structure,
        "reviewed": reviewed, "next_index": next_idx,
        "timestamp": NOW_ISO,
    }
    with open(session_path(family_key), "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, cls=NumpyEncoder)
    print(f"\n  [SAVED] Use --resume to continue: --family {family_key} --resume")


def load_session(family_key: str) -> Optional[dict]:
    p = session_path(family_key)
    if not p.exists():
        return None
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)


def interactive_review(classified: list[dict], family_key: str,
                       filename: str, metadata: dict, structure: dict,
                       start_idx: int = 0,
                       prior_reviewed: Optional[list] = None) -> list[dict]:
    """
    Review loop. Returns accepted rows.
    Handles Y/N/E/A/S/Q.
    """
    accepted = []
    skipped  = 0
    reviewed = prior_reviewed or []
    total    = len(classified)

    # Restore accepted from prior session
    for r in reviewed:
        if r.get("decision") == "accept":
            accepted.append(r["row"])

    idx = start_idx
    while idx < total:
        row = classified[idx]
        _print_review_card(row, idx, total)

        try:
            choice = input("> ").strip().upper()
        except (EOFError, KeyboardInterrupt):
            choice = "Q"

        if choice in ("Y", ""):
            accepted.append(row)
            reviewed.append({"index": idx, "decision": "accept", "row": row})
            idx += 1

        elif choice == "N":
            reviewed.append({"index": idx, "decision": "skip"})
            skipped += 1
            idx += 1

        elif choice == "E":
            row = _edit_row(row)
            accepted.append(row)
            reviewed.append({"index": idx, "decision": "accept", "row": row})
            idx += 1

        elif choice == "A":
            for rem in classified[idx:]:
                accepted.append(rem)
            print(f"\n  [INFO] Accepted all remaining {total - idx} pairs.")
            idx = total

        elif choice == "S":
            _print_stats(len(accepted), skipped, total)

        elif choice == "Q":
            save_session(family_key, filename, reviewed, idx, metadata, structure)
            return accepted

    return accepted


# ============================================================
# SECTION 11 — IMPROVE MODE (Planning)
# ============================================================
def embed_texts(texts: list[str]) -> list:
    from chromadb.utils import embedding_functions
    ef = embedding_functions.SentenceTransformerEmbeddingFunction(
        model_name="BAAI/bge-large-en-v1.5"
    )
    return ef(texts)


def cosine_sim(a: list, b: list) -> float:
    import math
    dot   = sum(x * y for x, y in zip(a, b))
    mag_a = math.sqrt(sum(x * x for x in a))
    mag_b = math.sqrt(sum(y * y for y in b))
    return dot / (mag_a * mag_b) if mag_a and mag_b else 0.0


def find_best_match(q_emb: list, embs: list,
                    threshold: float = 0.80) -> tuple[int, float]:
    best_i, best_s = -1, 0.0
    for i, emb in enumerate(embs):
        s = cosine_sim(q_emb, emb)
        if s > best_s:
            best_s, best_i = s, i
    return (best_i, best_s) if best_s >= threshold else (-1, best_s)


def interactive_review_improve(classified: list[dict], existing: list[dict],
                                existing_embs: list, family_key: str,
                                filename: str, metadata: dict,
                                structure: dict) -> tuple[list, list]:
    """IMPROVE mode review for Planning. Returns (new_accepted, improvements)."""
    print(f"\n[INFO] Embedding {len(classified)} historical questions...")
    hist_embs   = embed_texts([r["question"] for r in classified])
    new_entries = []
    improvements= []
    total       = len(classified)

    for idx, (row, h_emb) in enumerate(zip(classified, hist_embs)):
        best_i, best_s = find_best_match(h_emb, existing_embs)

        if best_i == -1:
            # New entry — regular review card
            _print_review_card(row, idx, total)
            try:
                choice = input("> ").strip().upper()
            except (EOFError, KeyboardInterrupt):
                save_session(family_key, filename, [], idx, metadata, structure)
                return new_entries, improvements

            if choice in ("Y", ""):
                new_entries.append(row)
            elif choice == "E":
                new_entries.append(_edit_row(row))
            elif choice == "Q":
                save_session(family_key, filename, [], idx, metadata, structure)
                return new_entries, improvements
        else:
            # Match found
            exist = existing[best_i]
            exist_q = exist.get("canonical_question", "") or exist.get("question", "")
            exist_a = exist.get("canonical_answer", "")  or exist.get("answer", "")

            print(f"\n{'='*60}")
            print(f"  IMPROVE | {idx+1}/{total} | Similarity: {best_s:.2f}")
            print(f"{'='*60}")
            print(f"\n  Q (new): ")
            print(_wrap(row["question"]))
            print(f"\n  A (new):")
            print(_wrap(row["answer"][:400]))
            print(f"\n  EXISTING ({exist.get('kb_id','') or exist.get('id','?')}):")
            print(_wrap(exist_q[:120]))
            print(_wrap(exist_a[:400]))
            print()
            print("-" * 60)
            print("  [K] Keep existing   [R] Replace   [M] Merge (edit)")
            print("  [N] Not a match (add as new)   [Q] Quit & save")
            print("-" * 60)

            try:
                choice = input("> ").strip().upper()
            except (EOFError, KeyboardInterrupt):
                choice = "Q"

            if choice == "K":
                pass
            elif choice == "R":
                improvements.append({
                    "action": "replace",
                    "existing_kb_id": exist.get("kb_id") or exist.get("id"),
                    "similarity": round(best_s, 4),
                    "new_answer": row["answer"],
                    "source": row.get("source_file", ""),
                })
            elif choice == "M":
                row = _edit_row(row)
                improvements.append({
                    "action": "merge",
                    "existing_kb_id": exist.get("kb_id") or exist.get("id"),
                    "similarity": round(best_s, 4),
                    "merged_answer": row["answer"],
                    "source": row.get("source_file", ""),
                })
            elif choice == "N":
                new_entries.append(row)
            elif choice == "Q":
                save_session(family_key, filename, [], idx, metadata, structure)
                return new_entries, improvements

    return new_entries, improvements


def auto_review_improve(classified: list[dict], existing: list[dict],
                        existing_embs: list, family_key: str,
                        filename: str, metadata: dict,
                        structure: dict, model: str) -> tuple[list, list]:
    """Auto IMPROVE mode using 5-stage answer selection. Returns (new_accepted, improvements)."""
    print(f"\n[INFO] Embedding {len(classified)} historical questions...")
    hist_embs   = embed_texts([r["question"] for r in classified])
    new_entries  = []
    improvements = []
    audit_log    = []
    decisions    = {"KEEP_EXISTING": 0, "REPLACE": 0, "ADD_NEW": 0}
    llm_budget   = LLM_BUDGET_PER_FILE
    total        = len(classified)
    client_name  = metadata.get("client", "")

    def _llm_call(prompt: str) -> str:
        return call_llm(prompt, model)

    for idx, (row, h_emb) in enumerate(zip(classified, hist_embs)):
        best_i, best_s = find_best_match(h_emb, existing_embs)

        if best_i == -1:
            # No match in KB — add as new
            new_entries.append(row)
            entry = {
                "index": idx,
                "decision": "ADD_NEW",
                "reason": f"No KB match (best sim: {best_s:.2f} < 0.80)",
                "stage": "no_match",
                "similarity": round(best_s, 4),
            }
            audit_log.append(entry)
            decisions["ADD_NEW"] += 1
            continue

        exist = existing[best_i]
        exist_q = exist.get("canonical_question", "") or exist.get("question", "")
        exist_a = exist.get("canonical_answer", "")  or exist.get("answer", "")
        exist_date = exist.get("last_updated", "") or exist.get("created_date", "")

        result = select_answer(
            existing_question=exist_q,
            existing_answer=exist_a,
            new_question=row["question"],
            new_answer=row["answer"],
            similarity=best_s,
            llm_call=_llm_call,
            client_name=client_name,
            existing_date=exist_date,
            new_date=metadata.get("date_estimated", ""),
            llm_calls_remaining=llm_budget,
        )

        if result.get("llm_used"):
            llm_budget -= 1

        decision = result["decision"]
        decisions[decision] = decisions.get(decision, 0) + 1

        log_entry = {
            "index": idx,
            "existing_kb_id": exist.get("kb_id") or exist.get("id", "?"),
            "similarity": round(best_s, 4),
            "decision": decision,
            "reason": result.get("reason", ""),
            "stage": result.get("stage", ""),
            "scores": result.get("scores"),
            "llm_used": result.get("llm_used", False),
        }
        audit_log.append(log_entry)

        if decision == "ADD_NEW":
            new_entries.append(row)
        elif decision == "REPLACE":
            improvements.append({
                "action": "replace",
                "existing_kb_id": exist.get("kb_id") or exist.get("id"),
                "similarity": round(best_s, 4),
                "new_answer": row["answer"],
                "source": row.get("source_file", filename),
                "reason": result.get("reason", ""),
                "stage": result.get("stage", ""),
            })
        # KEEP_EXISTING — no action needed

        # Progress every 25 entries
        if (idx + 1) % 25 == 0 or idx == total - 1:
            print(f"  [{idx+1}/{total}] kept={decisions['KEEP_EXISTING']} "
                  f"replace={decisions['REPLACE']} add={decisions['ADD_NEW']} "
                  f"llm_budget={llm_budget}")

    # Report
    llm_used = LLM_BUDGET_PER_FILE - llm_budget
    print_improve_report(decisions, audit_log, llm_used)

    report_path = save_improve_report(audit_log, filename)
    print(f"  [OK] Full audit log -> {report_path.name}")

    return new_entries, improvements


# ============================================================
# SECTION 12 — STAGE 4: ENTRY BUILDER (CLEAN V2)
# ============================================================
def build_v2_entry(row: dict, family_key: str, family: dict,
                   archive_id: str, seq: int) -> dict:
    """
    Build a canonical v2 entry with EXACTLY the fields in canonical_entry_v2.json.
    Uses question_generic if available, falls back to question.
    Never produces an empty answer.
    """
    cat      = row.get("category", "functional")
    cat_code = CATEGORY_CODES.get(cat, "GEN")
    entry_id = f"{family['id_prefix']}-{cat_code}-{seq:04d}"

    question = (row.get("question_generic") or row["question"]).strip()
    answer   = row["answer"].strip()

    # Sanity: should never reach here with empty answer, but guard anyway
    if not answer:
        raise ValueError(f"build_v2_entry: empty answer for row {row.get('row_num')}")

    return {
        "id":               entry_id,
        "question":         question,
        "answer":           answer,
        "question_variants":row.get("question_variants", []),
        "solution_codes":   row.get("solution_codes", []),
        "family_code":      family_key,
        "category":         cat,
        "subcategory":      row.get("subcategory", ""),
        "tags":             row.get("tags", []),
        "confidence":       "draft",
        "source_rfps":      [archive_id],
        "last_updated":     TODAY,
        "cloud_native_only":family.get("cloud_native", True),
        "notes":            "",
    }


# ============================================================
# SECTION 13 — ARCHIVE
# ============================================================
def archive_file(proc_path: Path, archived_name: str, metadata: dict,
                 structure: dict, all_rows: list[dict], accepted: list[dict],
                 archive_id: str, family_key: str) -> None:
    ARCHIVE_DIR.mkdir(exist_ok=True)
    (ARCHIVE_DIR / "files").mkdir(exist_ok=True)
    (ARCHIVE_DIR / "extractions").mkdir(exist_ok=True)

    stem = Path(archived_name).stem

    # Move original file
    shutil.move(str(proc_path), str(ARCHIVE_DIR / "files" / archived_name))

    # Save structure
    struct_path = ARCHIVE_DIR / "extractions" / f"{stem}_structure.json"
    with open(struct_path, "w", encoding="utf-8") as f:
        json.dump(structure, f, indent=2, ensure_ascii=False, cls=NumpyEncoder)

    # Save all extracted rows (before filtering) as JSONL
    ext_path = ARCHIVE_DIR / "extractions" / f"{stem}_extracted.jsonl"
    with open(ext_path, "w", encoding="utf-8") as f:
        for r in all_rows:
            f.write(json.dumps(r, ensure_ascii=False, cls=NumpyEncoder) + "\n")

    # Update registry
    reg         = load_registry()
    cat_counts  = Counter(r.get("category", "general") for r in accepted)
    n_processed = sum(1 for sh in structure.get("relevant_sheets", [])
                      if sh.get("purpose") != "skip")

    entry = {
        "archive_id":        archive_id,
        "original_filename": proc_path.name,
        "archived_filename": archived_name,
        "client":            metadata.get("client", ""),
        "client_industry":   metadata.get("client_industry", ""),
        "family_code":       family_key,
        "solution_codes":    [],
        "rfp_type":          metadata.get("rfp_type", "response"),
        "date_estimated":    metadata.get("date_estimated", ""),
        "date_processed":    TODAY,
        "region":            metadata.get("region", ""),
        "extraction_stats": {
            "total_sheets":       len(structure.get("relevant_sheets", [])),
            "sheets_processed":   n_processed,
            "total_qa_extracted": len(all_rows),
            "accepted":           len(accepted),
            "skipped":            len(all_rows) - len(accepted),
            "categories":         dict(cat_counts),
        },
        "canonical_entries_added": len(accepted),
        "structure_file":    f"extractions/{stem}_structure.json",
        "extraction_file":   f"extractions/{stem}_extracted.jsonl",
        "notes":             metadata.get("notes", ""),
        "tags":              [],
    }
    reg["entries"].append(entry)
    reg["total_files"]        = len(reg["entries"])
    reg["total_qa_extracted"] = sum(
        e["extraction_stats"]["accepted"] for e in reg["entries"]
    )
    save_registry(reg)
    print(f"[OK] Archived -> {archived_name}  ({archive_id})")


# ============================================================
# SECTION 14 — MAIN FILE FLOW
# ============================================================
def get_inbox_files(family_key: str, specific: Optional[str]) -> list[Path]:
    if specific:
        p = Path(specific)
        return [p] if p.exists() else []
    inbox = HISTORICAL_DIR / family_key / "inbox"
    files = list(inbox.glob("*.xlsx")) + list(inbox.glob("*.xls"))
    return sorted(f for f in files if not f.name.startswith("~$"))


def process_file(filepath: Path, family_key: str, family: dict,
                 model: str, is_improve: bool,
                 existing: list, existing_embs: list,
                 resume_session: Optional[dict],
                 interactive: bool = False,
                 batch_mode: bool = False) -> Optional[dict]:
    """Full pipeline for one file."""
    proc_dir = HISTORICAL_DIR / family_key / "processing"
    proc_dir.mkdir(exist_ok=True)

    # ---- Resume? ----
    if resume_session and resume_session.get("file") == filepath.name:
        print(f"\n[RESUME] {filepath.name} from index {resume_session['next_index']}")
        metadata  = resume_session["metadata"]
        structure = resume_session["structure"]
        proc_path = proc_dir / filepath.name
        if not proc_path.exists():
            shutil.copy(str(filepath), str(proc_path))
        start_idx      = resume_session["next_index"]
        prior_reviewed = resume_session.get("reviewed", [])
    else:
        # STAGE 1 — Structure
        print(f"\n[STAGE 1] Scanning: {filepath.name}")
        try:
            prescan = prescan_excel(filepath)
        except Exception as e:
            print(f"  [ERROR] Cannot open file: {e}")
            return None

        metadata = collect_metadata_interactive(filepath.name, family_key)

        print(f"\n[INFO] Analyzing structure with {model.upper()}...")
        try:
            structure = analyze_structure_llm(prescan, family["display_name"], model)
        except Exception as e:
            print(f"  [ERROR] Structure analysis failed: {e}")
            return None

        if not confirm_structure(structure, filepath.name):
            print(f"  [SKIP] {filepath.name}")
            return None

        proc_path = proc_dir / filepath.name
        shutil.copy(str(filepath), str(proc_path))
        start_idx      = 0
        prior_reviewed = []

    # STAGE 2 — Extract + Filter
    print(f"\n[STAGE 2] Extracting rows from: {proc_path.name}")
    try:
        all_rows = extract_all_rows(proc_path, structure)
    except Exception as e:
        print(f"  [ERROR] Extraction failed: {e}")
        shutil.copy(str(proc_path), str(filepath))
        proc_path.unlink(missing_ok=True)
        return None

    print(f"  Total rows extracted: {len(all_rows)}")

    candidates, prog_stats = programmatic_filter(all_rows)
    print(f"  After programmatic filter: {len(candidates)} candidates")

    if not candidates:
        print("  [WARN] No candidates after programmatic filter.")
        print(f"         Source RFP with no BY answers? "
              f"({prog_stats.get('skip_empty_answer',0)} rows had empty answers)")
        return None

    print(f"\n  LLM content filter ({len(candidates)} candidates, batch=10)...")
    kept, llm_counts = llm_content_filter(candidates, family["display_name"], model)

    print_filter_stats(len(all_rows), prog_stats, llm_counts, len(kept))

    if not kept:
        print("  [WARN] Nothing kept after content filtering.")
        return None

    ans = input(f"\n  {len(kept)} rows to review. Continue? [Y/n]: ").strip().lower()
    if ans not in ("", "y", "yes"):
        return None

    # STAGE 3 — Anonymize + Classify + Review
    client_name = metadata.get("client", "")
    if client_name and client_name.lower() != "unknown":
        kept = anonymize_rows(kept, client_name)
        print(f"\n  Anonymized '{client_name}' -> [Customer] in questions + answers")

    print(f"\n[STAGE 3] Classifying {len(kept)} rows with {model.upper()}...")
    if batch_mode:
        classified = classify_rows_batch_async(kept, family, model)
    else:
        classified = classify_rows_batch(kept, family, model)

    if is_improve and not interactive:
        print(f"\n[STAGE 3] Auto-review IMPROVE ({len(classified)} entries)...")
        accepted, improvements = auto_review_improve(
            classified, existing, existing_embs,
            family_key, filepath.name, metadata, structure, model,
        )
    elif is_improve:
        print(f"\n[STAGE 3] Interactive review ({len(classified)} entries)...")
        accepted, improvements = interactive_review_improve(
            classified, existing, existing_embs,
            family_key, filepath.name, metadata, structure,
        )
    else:
        print(f"\n[STAGE 3] Interactive review ({len(classified)} entries)...")
        accepted = interactive_review(
            classified, family_key, filepath.name, metadata, structure,
            start_idx, prior_reviewed,
        )
        improvements = []

    if not accepted:
        print("  [INFO] No entries accepted.")
        return None

    print(f"\n  Accepted: {len(accepted)} entries.")

    # STAGE 4 — Build + Save + Archive
    existing_canon = load_canonical(family["canonical_file"])
    seq_start      = next_seq(existing_canon, family["id_prefix"])
    reg            = load_registry()
    arc_id         = next_archive_id(reg)

    new_entries = []
    for i, row in enumerate(accepted):
        try:
            entry = build_v2_entry(row, family_key, family, arc_id, seq_start + i)
            new_entries.append(entry)
        except ValueError as e:
            print(f"  [WARN] Skipping entry {i}: {e}")

    if not new_entries:
        print("  [WARN] No valid entries after build step.")
        return None

    combined = existing_canon + new_entries
    save_canonical(family["canonical_file"], combined)

    # Staging: improvements (IMPROVE mode)
    if improvements:
        STAGING_DIR.mkdir(exist_ok=True)
        imp_path = STAGING_DIR / f"{family_key}_improvements.jsonl"
        with open(imp_path, "a", encoding="utf-8") as f:
            for item in improvements:
                f.write(json.dumps(item, ensure_ascii=False, cls=NumpyEncoder) + "\n")
        print(f"  {len(improvements)} improvement candidates -> {imp_path.name}")

    # Archive
    archived_name = make_archived_filename(metadata, family_key, filepath.suffix)
    archive_file(proc_path, archived_name, metadata, structure,
                 all_rows, accepted, arc_id, family_key)

    # Cleanup
    filepath.unlink(missing_ok=True)            # remove from inbox
    session_path(family_key).unlink(missing_ok=True)  # clear session

    return {"accepted": len(new_entries), "archive_id": arc_id}


def run_family(family_key: str, model: str, resume: bool,
               specific_file: Optional[str], interactive: bool = False,
               batch_mode: bool = False) -> None:
    family     = get_family(family_key)
    is_improve = family.get("phase", 1) == 2

    print(f"\n[START] {family['display_name']}  model={model}  "
          f"mode={'IMPROVE' if is_improve else 'CREATE'}")

    resume_session = load_session(family_key) if resume else None
    if resume and not resume_session:
        print(f"  [WARN] No saved session found for '{family_key}'.")

    existing, existing_embs = [], []
    if is_improve:
        existing = load_canonical(family["canonical_file"])
        if existing:
            print(f"[INFO] Pre-embedding {len(existing)} existing entries (~30s)...")
            q_field      = lambda e: e.get("canonical_question","") or e.get("question","")
            existing_embs = embed_texts([q_field(e) for e in existing])
            print("[OK]  Embeddings ready.")

    if specific_file:
        files = [Path(specific_file)]
    elif resume_session:
        for candidate in [
            HISTORICAL_DIR / family_key / "processing" / resume_session["file"],
            HISTORICAL_DIR / family_key / "inbox"      / resume_session["file"],
        ]:
            if candidate.exists():
                files = [candidate]
                break
        else:
            files = []
    else:
        files = get_inbox_files(family_key, None)

    if not files:
        print(f"\n[INFO] No files in historical/{family_key}/inbox/")
        print(f"       Drop .xlsx files there and re-run.")
        return

    print(f"[INFO] {len(files)} file(s) to process")

    total = 0
    for f in files:
        result = process_file(f, family_key, family, model,
                              is_improve, existing, existing_embs, resume_session,
                              interactive=interactive, batch_mode=batch_mode)
        if result:
            total += result["accepted"]
            resume_session = None

    if total:
        print(f"\n[DONE] {total} new entries added.")
        print("[INFO] Syncing ChromaDB index...")
        try:
            from kb_index_sync import sync
            sync()
        except Exception as e:
            print(f"[WARN] Auto-sync failed: {e}")
            print("       Run manually: python src/kb_index_sync.py")


# ============================================================
# SECTION 15 — ENTRY POINT
# ============================================================
def main() -> None:
    import argparse
    FAMILIES = ["planning","wms","logistics","scpo","catman","workforce",
                "commerce","flexis","network","doddle","aiml"]

    p = argparse.ArgumentParser(
        description="3-Stage RFP Extraction Pipeline (row = atomic unit)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python src/kb_extract_historical.py --family network
  python src/kb_extract_historical.py --family network --model gemini-flash
  python src/kb_extract_historical.py --family planning          # auto IMPROVE
  python src/kb_extract_historical.py --family wms --resume
  python src/kb_extract_historical.py --family wms --file "path/to.xlsx"
        """,
    )
    p.add_argument("--family",  required=True, choices=FAMILIES)
    p.add_argument("--model",   default="gemini-flash",
                   help="LLM model key (default: gemini-flash)")
    p.add_argument("--resume",  action="store_true",
                   help="Resume an interrupted review session")
    p.add_argument("--file",    default=None,
                   help="Process a specific Excel file instead of scanning inbox/")
    p.add_argument("--interactive", action="store_true",
                   help="Use interactive review for IMPROVE mode (default: auto-mode)")
    p.add_argument("--batch", action="store_true",
                   help="Use Gemini Batch API for classification (50%% cheaper, async)")
    args = p.parse_args()

    run_family(args.family, args.model, args.resume, args.file,
               interactive=args.interactive, batch_mode=args.batch)


if __name__ == "__main__":
    main()
