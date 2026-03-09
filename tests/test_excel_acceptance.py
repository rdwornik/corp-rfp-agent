"""Acceptance tests for Excel agent -- structure checks only (no LLM calls)."""

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

# Add src/ to import path
PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from rfp_excel_agent import (
    is_green_cell,
    find_header_row,
    detect_question_column,
    detect_answer_column,
    scan_green_cells,
)


# ---------------------------------------------------------------------------
# Test 1: Green cells are correctly identified
# ---------------------------------------------------------------------------
def test_green_cells_detected(excel_golden):
    """Agent finds exactly the green (unanswered) cells."""
    wb = load_workbook(str(excel_golden))
    ws = wb["Requirements"]

    green_rows = []
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if is_green_cell(cell):
                green_rows.append(row_idx)
                break

    # Rows 2, 3, 5, 7, 8, 9 have green cells
    assert sorted(green_rows) == [2, 3, 5, 7, 8, 9]


# ---------------------------------------------------------------------------
# Test 2: Already-answered cells are NOT green
# ---------------------------------------------------------------------------
def test_answered_cells_not_green(excel_golden):
    """Cells with existing answers are not detected as green."""
    wb = load_workbook(str(excel_golden))
    ws = wb["Requirements"]

    # Row 4 and 6 have answers, should not have green cells
    for row_idx in [4, 6]:
        has_green = False
        for col_idx in range(1, ws.max_column + 1):
            if is_green_cell(ws.cell(row=row_idx, column=col_idx)):
                has_green = True
                break
        assert not has_green, f"Row {row_idx} should NOT have green cells"


# ---------------------------------------------------------------------------
# Test 3: Header row detected correctly
# ---------------------------------------------------------------------------
def test_header_row_detection(excel_golden):
    """Header row is detected at row 1."""
    wb = load_workbook(str(excel_golden))
    ws = wb["Requirements"]

    header_row = find_header_row(ws)
    assert header_row == 1


# ---------------------------------------------------------------------------
# Test 4: Question column detected correctly
# ---------------------------------------------------------------------------
def test_question_column_detection(excel_golden):
    """Question column is detected as column C (3) via 'Requirement' header."""
    wb = load_workbook(str(excel_golden))
    ws = wb["Requirements"]

    header_row = find_header_row(ws)
    col_idx, col_name = detect_question_column(ws, header_row)

    assert col_idx == 3  # Column C
    assert "Requirement" in col_name


# ---------------------------------------------------------------------------
# Test 5: Answer column detected correctly
# ---------------------------------------------------------------------------
def test_answer_column_detection(excel_golden):
    """Answer column is detected as column E (5) via 'Vendor Response' header."""
    wb = load_workbook(str(excel_golden))
    ws = wb["Requirements"]

    header_row = find_header_row(ws)
    question_col, _ = detect_question_column(ws, header_row)
    answer_col, answer_name = detect_answer_column(ws, header_row, question_col)

    assert answer_col == 5  # Column E
    assert "response" in answer_name.lower() or "vendor" in answer_name.lower()


# ---------------------------------------------------------------------------
# Test 6: scan_green_cells returns correct rows
# ---------------------------------------------------------------------------
def test_scan_green_cells(excel_golden):
    """scan_green_cells finds all green rows with question text."""
    wb = load_workbook(str(excel_golden))

    results = scan_green_cells(wb)
    found_rows = [r["row"] for r in results]

    # Should find rows 2, 3, 5, 7, 8, 9
    assert 2 in found_rows
    assert 3 in found_rows
    assert 5 in found_rows
    assert 7 in found_rows
    assert 8 in found_rows
    assert 9 in found_rows

    # Should NOT include row 4 (answered), 6 (answered), 10 (header)
    assert 4 not in found_rows
    assert 6 not in found_rows
    assert 10 not in found_rows


# ---------------------------------------------------------------------------
# Test 7: Merged cells handled correctly
# ---------------------------------------------------------------------------
def test_merged_cells(excel_golden):
    """Agent handles merged cells without error and extracts question text."""
    wb = load_workbook(str(excel_golden))
    results = scan_green_cells(wb)

    row7_results = [r for r in results if r["row"] == 7]
    assert len(row7_results) == 1
    assert "concurrent users" in row7_results[0]["question_text"]


# ---------------------------------------------------------------------------
# Test 8: Formula preserved
# ---------------------------------------------------------------------------
def test_formula_preservation(excel_golden):
    """Formula in adjacent cell F8 is present in the fixture."""
    wb = load_workbook(str(excel_golden))
    ws = wb["Requirements"]

    cell_f8 = ws["F8"]
    assert cell_f8.value == "=LEN(E8)"


# ---------------------------------------------------------------------------
# Test 9: Workbook integrity
# ---------------------------------------------------------------------------
def test_workbook_integrity(excel_golden):
    """Golden fixture opens without corruption."""
    wb = load_workbook(str(excel_golden))

    assert "Requirements" in wb.sheetnames
    ws = wb["Requirements"]
    assert ws.max_row >= 10
    assert ws.max_column >= 6


# ---------------------------------------------------------------------------
# Test 10: Category headers skipped
# ---------------------------------------------------------------------------
def test_category_headers_skipped(excel_golden):
    """Bold header row (row 10) is not returned by scan_green_cells."""
    wb = load_workbook(str(excel_golden))
    results = scan_green_cells(wb)

    found_rows = [r["row"] for r in results]
    assert 10 not in found_rows


# ---------------------------------------------------------------------------
# Test 11: Special characters preserved
# ---------------------------------------------------------------------------
def test_special_characters(excel_golden):
    """Unicode and special chars in questions survive round-trip."""
    wb = load_workbook(str(excel_golden))
    results = scan_green_cells(wb)

    row9_results = [r for r in results if r["row"] == 9]
    assert len(row9_results) == 1

    question = row9_results[0]["question_text"]
    assert '"quotes"' in question
    assert "&" in question
    assert "\u0142\u00f3\u015b\u0107" in question
