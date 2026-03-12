"""Tests for generate_review_pack -- Excel review pack generation."""

import json
import sys
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from generate_review_pack import (
    load_profile,
    load_existing_questions,
    generate_questions,
    _parse_questions,
    answer_question,
    _empty_answer,
    _content_hash,
    create_review_pack,
    generate_pack,
    QUESTION_TOPICS,
    BY_MIDNIGHT_BLUE,
    YELLOW_BG,
    GENERATOR_VERSION,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

SAMPLE_PROFILE = {
    "display_name": "WMS",
    "product": "wms",
    "cloud_native": True,
    "deployment": ["SaaS", "Hybrid"],
    "apis": ["REST", "GraphQL"],
    "key_facts": ["Real-time inventory", "Android mobile app"],
    "forbidden_claims": ["On-premise only", "No API support"],
    "_meta": {"status": "active"},
}


def _make_question(q="What APIs?", topic="integration", cat="technical", src="profile"):
    return {"question": q, "topic": topic, "category": cat, "source": src}


def _make_rag_result(answer="REST API", confidence=0.85, entry_ids=None):
    if entry_ids is None:
        entry_ids = ["WMS-FUNC-0001"]
    return {
        "answer": answer,
        "entry_ids": entry_ids,
        "confidence": confidence,
        "revision_hash": _content_hash(answer) if answer else "",
    }


# ---------------------------------------------------------------------------
# Question generation
# ---------------------------------------------------------------------------

class TestParseQuestions:
    def test_parse_valid_json_array(self):
        raw = json.dumps([{"question": "Q1"}, {"question": "Q2"}])
        result = _parse_questions(raw)
        assert len(result) == 2
        assert result[0]["question"] == "Q1"

    def test_parse_markdown_fenced(self):
        raw = '```json\n[{"question": "Q1"}]\n```'
        result = _parse_questions(raw)
        assert len(result) == 1

    def test_parse_empty(self):
        assert _parse_questions("") == []
        assert _parse_questions(None) == []

    def test_parse_filters_invalid_entries(self):
        raw = json.dumps([{"question": "valid"}, {"no_question_key": "bad"}])
        result = _parse_questions(raw)
        assert len(result) == 1


class TestGenerateQuestions:
    @patch("kb_extract_historical.call_llm")
    def test_correct_count(self, mock_llm):
        questions = [{"question": f"Q{i}", "topic": "security", "category": "technical", "source": "profile"}
                     for i in range(10)]
        mock_llm.return_value = json.dumps(questions)
        result = generate_questions(SAMPLE_PROFILE, count=10, model="gemini-flash")
        assert len(result) <= 10

    @patch("kb_extract_historical.call_llm")
    def test_includes_existing(self, mock_llm):
        llm_qs = [{"question": f"LLM-Q{i}"} for i in range(5)]
        mock_llm.return_value = json.dumps(llm_qs)
        existing = [_make_question(q=f"Existing-{i}", src="existing") for i in range(10)]
        result = generate_questions(SAMPLE_PROFILE, count=10, existing=existing, model="gemini-flash")
        sources = [q.get("source") for q in result]
        assert "existing" in sources

    @patch("kb_extract_historical.call_llm")
    def test_required_fields(self, mock_llm):
        mock_llm.return_value = json.dumps([{"question": "Q1"}])
        result = generate_questions(SAMPLE_PROFILE, count=5, model="gemini-flash")
        for q in result:
            assert "question" in q
            assert "topic" in q
            assert "category" in q
            assert "source" in q


# ---------------------------------------------------------------------------
# RAG answering
# ---------------------------------------------------------------------------

class TestAnswerQuestion:
    def test_empty_answer_when_no_chroma(self, tmp_path):
        result = answer_question("Q?", "wms", chroma_path=tmp_path / "nonexistent")
        assert result == _empty_answer()
        assert result["answer"] == ""
        assert result["confidence"] == 0.0

    @patch.dict("sys.modules", {
        "chromadb": MagicMock(),
        "chromadb.utils": MagicMock(),
        "chromadb.utils.embedding_functions": MagicMock(),
    })
    def test_returns_answer_on_match(self, tmp_path):
        chroma_dir = tmp_path / "chroma"
        chroma_dir.mkdir()

        mock_collection = MagicMock()
        mock_collection.query.return_value = {
            "ids": [["wms_0001"]],
            "distances": [[0.3]],
            "metadatas": [[{"canonical_answer": "REST API supported", "kb_id": "WMS-0001"}]],
        }

        mock_client = MagicMock()
        mock_client.get_collection.return_value = mock_collection

        chromadb_mod = sys.modules["chromadb"]
        chromadb_mod.PersistentClient.return_value = mock_client

        result = answer_question("What APIs?", "wms", chroma_path=chroma_dir)
        assert result["answer"] == "REST API supported"
        assert result["confidence"] > 0
        assert "WMS-0001" in result["entry_ids"]
        assert result["revision_hash"] != ""

    def test_confidence_calculated(self):
        # Test content hash utility
        h = _content_hash("test answer")
        assert len(h) == 16
        assert h == _content_hash("test answer")  # deterministic


# ---------------------------------------------------------------------------
# Excel creation
# ---------------------------------------------------------------------------

class TestCreateReviewPack:
    def test_creates_xlsx(self, tmp_path):
        questions = [_make_question()]
        rag_results = [_make_rag_result()]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        assert path.exists()
        assert path.suffix == ".xlsx"

    def test_three_sheets(self, tmp_path):
        from openpyxl import load_workbook
        questions = [_make_question()]
        rag_results = [_make_rag_result()]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        wb = load_workbook(str(path))
        assert "Questions & Answers" in wb.sheetnames
        assert "General Feedback" in wb.sheetnames
        assert "_metadata" in wb.sheetnames
        wb.close()

    def test_header_formatting(self, tmp_path):
        from openpyxl import load_workbook
        questions = [_make_question()]
        rag_results = [_make_rag_result()]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        wb = load_workbook(str(path))
        ws = wb["Questions & Answers"]
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.fill.start_color.rgb == "00" + BY_MIDNIGHT_BLUE
        assert header_cell.font.bold
        wb.close()

    def test_dropdown_validation(self, tmp_path):
        from openpyxl import load_workbook
        questions = [_make_question(), _make_question(q="Q2")]
        rag_results = [_make_rag_result(), _make_rag_result()]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        wb = load_workbook(str(path))
        ws = wb["Questions & Answers"]
        # DataValidation should be present
        assert len(ws.data_validations.dataValidation) > 0
        dv = ws.data_validations.dataValidation[0]
        assert "APPROVE" in dv.formula1
        assert "REJECT" in dv.formula1
        wb.close()

    def test_prefilled_action_approve(self, tmp_path):
        from openpyxl import load_workbook
        questions = [_make_question()]
        rag_results = [_make_rag_result(confidence=0.9)]  # >= 0.8 → APPROVE
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        wb = load_workbook(str(path))
        ws = wb["Questions & Answers"]
        action_cell = ws.cell(row=2, column=8)
        assert action_cell.value == "APPROVE"
        wb.close()

    def test_prefilled_action_new_when_empty(self, tmp_path):
        from openpyxl import load_workbook
        questions = [_make_question()]
        rag_results = [_make_rag_result(answer="", confidence=0.0)]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        wb = load_workbook(str(path))
        ws = wb["Questions & Answers"]
        action_cell = ws.cell(row=2, column=8)
        assert action_cell.value == "NEW"
        wb.close()

    def test_prefilled_action_skip_when_low_confidence(self, tmp_path):
        from openpyxl import load_workbook
        questions = [_make_question()]
        rag_results = [_make_rag_result(confidence=0.5)]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        wb = load_workbook(str(path))
        ws = wb["Questions & Answers"]
        action_cell = ws.cell(row=2, column=8)
        assert action_cell.value == "SKIP"
        wb.close()

    def test_metadata_hidden(self, tmp_path):
        from openpyxl import load_workbook
        questions = [_make_question()]
        rag_results = [_make_rag_result()]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        wb = load_workbook(str(path))
        assert wb["_metadata"].sheet_state == "hidden"
        wb.close()

    def test_hidden_columns_k_l(self, tmp_path):
        from openpyxl import load_workbook
        questions = [_make_question()]
        rag_results = [_make_rag_result()]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        wb = load_workbook(str(path))
        ws = wb["Questions & Answers"]
        assert ws.column_dimensions["K"].hidden
        assert ws.column_dimensions["L"].hidden
        wb.close()

    def test_confidence_color_green(self, tmp_path):
        from openpyxl import load_workbook
        questions = [_make_question()]
        rag_results = [_make_rag_result(confidence=0.9)]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        wb = load_workbook(str(path))
        ws = wb["Questions & Answers"]
        conf_cell = ws.cell(row=2, column=7)
        assert conf_cell.font.color.rgb == "00006600"
        wb.close()

    def test_empty_answer_red_italic(self, tmp_path):
        from openpyxl import load_workbook
        questions = [_make_question()]
        rag_results = [_make_rag_result(answer="", confidence=0.0)]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        wb = load_workbook(str(path))
        ws = wb["Questions & Answers"]
        answer_cell = ws.cell(row=2, column=5)
        assert answer_cell.font.italic
        assert answer_cell.value == "No answer found in KB"
        wb.close()

    def test_feedback_columns_yellow(self, tmp_path):
        from openpyxl import load_workbook
        questions = [_make_question()]
        rag_results = [_make_rag_result()]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        wb = load_workbook(str(path))
        ws = wb["Questions & Answers"]
        feedback_cell = ws.cell(row=2, column=9)
        robs_cell = ws.cell(row=2, column=10)
        assert feedback_cell.fill.start_color.rgb == "00" + YELLOW_BG
        assert robs_cell.fill.start_color.rgb == "00" + YELLOW_BG
        wb.close()

    def test_freeze_panes(self, tmp_path):
        from openpyxl import load_workbook
        questions = [_make_question()]
        rag_results = [_make_rag_result()]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        wb = load_workbook(str(path))
        ws = wb["Questions & Answers"]
        assert ws.freeze_panes == "A2"
        wb.close()

    def test_metadata_content(self, tmp_path):
        from openpyxl import load_workbook
        questions = [_make_question()]
        rag_results = [_make_rag_result()]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        wb = load_workbook(str(path))
        ws = wb["_metadata"]
        meta = {}
        for row in ws.iter_rows(values_only=True):
            if row[0]:
                meta[row[0]] = row[1]
        assert meta["family"] == "wms"
        assert meta["generator_version"] == GENERATOR_VERSION
        assert "question_count" in meta
        wb.close()


# ---------------------------------------------------------------------------
# Pack naming and output directory
# ---------------------------------------------------------------------------

class TestPackNaming:
    def test_output_in_specified_dir(self, tmp_path):
        questions = [_make_question()]
        rag_results = [_make_rag_result()]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path / "packs")
        assert path.parent == tmp_path / "packs"
        assert path.exists()

    def test_filename_contains_family(self, tmp_path):
        questions = [_make_question()]
        rag_results = [_make_rag_result()]
        path = create_review_pack("wms", questions, rag_results, SAMPLE_PROFILE,
                                  output_dir=tmp_path)
        assert "WMS" in path.name
