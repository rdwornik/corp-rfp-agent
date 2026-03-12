"""Tests for import_review_pack -- Excel feedback import pipeline."""

import json
import sys
from datetime import datetime
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from import_review_pack import (
    parse_feedback_rules,
    parse_feedback_llm,
    check_stale,
    is_already_processed,
    mark_processed,
    validate_excel,
    parse_rows,
    parse_general_feedback,
    execute_row,
    _create_draft,
    import_pack,
    print_import_report,
    FEEDBACK_PATTERNS,
    INTENT_TO_ACTION,
    REQUIRED_HEADERS,
)


# ---------------------------------------------------------------------------
# Helper: create a minimal review pack Excel
# ---------------------------------------------------------------------------

def _create_pack_excel(path: Path, rows=None, metadata=None, general_fb=None):
    """Create a minimal Excel review pack for testing."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Questions & Answers"

    headers = ["#", "Question", "Topic", "Category", "Answer",
               "KB Entries Used", "Confidence", "Action",
               "Feedback", "Rob's Answer", "_revision_hash", "_entry_ids_json"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    if rows is None:
        rows = [_default_row()]

    for row_idx, row_data in enumerate(rows, 2):
        vals = [
            row_data.get("num", row_idx - 1),
            row_data.get("question", "What APIs?"),
            row_data.get("topic", "integration"),
            row_data.get("category", "technical"),
            row_data.get("answer", "REST API"),
            row_data.get("kb_used", "WMS-FUNC-0001"),
            row_data.get("confidence", 0.85),
            row_data.get("action", "APPROVE"),
            row_data.get("feedback", ""),
            row_data.get("robs_answer", ""),
            row_data.get("revision_hash", "abc123"),
            row_data.get("entry_ids_json", '["WMS-FUNC-0001"]'),
        ]
        for col_idx, val in enumerate(vals, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # General Feedback sheet
    ws2 = wb.create_sheet("General Feedback")
    ws2.cell(row=1, column=1, value="#")
    ws2.cell(row=1, column=2, value="Feedback")
    ws2.cell(row=1, column=3, value="Scope")
    if general_fb:
        for i, fb in enumerate(general_fb, 2):
            ws2.cell(row=i, column=1, value=i - 1)
            ws2.cell(row=i, column=2, value=fb.get("feedback", ""))
            ws2.cell(row=i, column=3, value=fb.get("scope", "general"))

    # _metadata sheet
    ws3 = wb.create_sheet("_metadata")
    meta = metadata or {"pack_id": "WMS_20260312_PACK001", "family": "wms",
                        "generated_at": "2026-03-12T10:00:00"}
    for i, (k, v) in enumerate(meta.items(), 1):
        ws3.cell(row=i, column=1, value=k)
        ws3.cell(row=i, column=2, value=str(v))

    wb.save(str(path))
    return path


def _default_row(**overrides):
    base = {
        "question": "What APIs does WMS support?",
        "answer": "REST API supported",
        "action": "APPROVE",
        "confidence": 0.85,
        "entry_ids_json": '["WMS-FUNC-0001"]',
        "revision_hash": "abc123",
    }
    base.update(overrides)
    return base


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

class TestValidation:
    def test_valid_excel(self, tmp_path):
        xlsx = _create_pack_excel(tmp_path / "pack.xlsx")
        ok, err, meta = validate_excel(xlsx)
        assert ok
        assert err == ""
        assert meta["pack_id"] == "WMS_20260312_PACK001"

    def test_missing_sheet(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.title = "WrongName"
        xlsx = tmp_path / "bad.xlsx"
        wb.save(str(xlsx))
        ok, err, _ = validate_excel(xlsx)
        assert not ok
        assert "Questions & Answers" in err

    def test_missing_column(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Questions & Answers"
        ws.cell(row=1, column=1, value="Question")
        # Missing "Answer" and "Action"
        xlsx = tmp_path / "missing_col.xlsx"
        wb.save(str(xlsx))
        ok, err, _ = validate_excel(xlsx)
        assert not ok
        assert "Answer" in err or "Action" in err

    def test_already_processed(self, tmp_path):
        packs_file = tmp_path / "processed.jsonl"
        packs_file.write_text(
            json.dumps({"pack_id": "PACK001"}) + "\n", encoding="utf-8"
        )
        with patch("import_review_pack.PROCESSED_PACKS", packs_file):
            assert is_already_processed("PACK001")
            assert not is_already_processed("PACK999")

    def test_force_allows_reprocess(self, tmp_path):
        """Force flag bypasses already-processed check in import_pack."""
        xlsx = _create_pack_excel(tmp_path / "pack.xlsx")
        packs_file = tmp_path / "processed.jsonl"
        packs_file.write_text(
            json.dumps({"pack_id": "WMS_20260312_PACK001"}) + "\n", encoding="utf-8"
        )
        with patch("import_review_pack.PROCESSED_PACKS", packs_file), \
             patch("import_review_pack.execute_row", return_value={"row": 2, "action": "APPROVE", "status": "WOULD_APPROVE"}), \
             patch("import_review_pack.parse_general_feedback", return_value=[]):
            summary = import_pack(xlsx, dry_run=True, force=True)
            assert summary["total"] > 0


# ---------------------------------------------------------------------------
# Row parsing
# ---------------------------------------------------------------------------

class TestRowParsing:
    def test_approve_action(self, tmp_path):
        xlsx = _create_pack_excel(tmp_path / "p.xlsx",
                                  rows=[_default_row(action="APPROVE")])
        rows = parse_rows(xlsx)
        assert rows[0]["action"] == "APPROVE"
        assert not rows[0]["skipped"]

    def test_update_action(self, tmp_path):
        xlsx = _create_pack_excel(tmp_path / "p.xlsx",
                                  rows=[_default_row(action="UPDATE")])
        rows = parse_rows(xlsx)
        assert rows[0]["action"] == "UPDATE"

    def test_reject_action(self, tmp_path):
        xlsx = _create_pack_excel(tmp_path / "p.xlsx",
                                  rows=[_default_row(action="REJECT")])
        rows = parse_rows(xlsx)
        assert rows[0]["action"] == "REJECT"

    def test_new_action(self, tmp_path):
        xlsx = _create_pack_excel(tmp_path / "p.xlsx",
                                  rows=[_default_row(action="NEW")])
        rows = parse_rows(xlsx)
        assert rows[0]["action"] == "NEW"

    def test_skip_action(self, tmp_path):
        xlsx = _create_pack_excel(tmp_path / "p.xlsx",
                                  rows=[_default_row(action="SKIP")])
        rows = parse_rows(xlsx)
        assert rows[0]["action"] == "SKIP"
        assert rows[0]["skipped"]

    def test_blank_action_skipped(self, tmp_path):
        xlsx = _create_pack_excel(tmp_path / "p.xlsx",
                                  rows=[_default_row(action="")])
        rows = parse_rows(xlsx)
        assert rows[0]["skipped"]


# ---------------------------------------------------------------------------
# Bilingual feedback parsing
# ---------------------------------------------------------------------------

class TestFeedbackParsing:
    def test_polish_brakuje(self):
        assert parse_feedback_rules("brakuje informacji") == "add_detail"

    def test_polish_zle(self):
        assert parse_feedback_rules("to jest zle") == "reject"

    def test_english_wrong(self):
        assert parse_feedback_rules("This is wrong") == "reject"

    def test_english_missing(self):
        assert parse_feedback_rules("Missing API details") == "add_detail"

    def test_english_ok(self):
        assert parse_feedback_rules("ok") == "approve"

    def test_english_good(self):
        assert parse_feedback_rules("Looks good") == "approve"

    def test_no_match(self):
        assert parse_feedback_rules("xxxxxx random text") is None

    def test_empty(self):
        assert parse_feedback_rules("") is None
        assert parse_feedback_rules(None) is None

    @patch("kb_extract_historical.call_llm")
    def test_llm_fallback(self, mock_llm):
        mock_llm.return_value = "UPDATE"
        result = parse_feedback_llm("Prosze poprawic odpowiedz", "What APIs?")
        assert result == "UPDATE"

    @patch("kb_extract_historical.call_llm")
    def test_llm_fallback_skip_on_garbage(self, mock_llm):
        mock_llm.return_value = "some garbage response"
        result = parse_feedback_llm("xyz", "Q?")
        assert result == "SKIP"

    def test_robs_answer_priority(self, tmp_path):
        """When Rob's Answer is provided, UPDATE should use it directly."""
        row = {
            "row": 2, "action": "UPDATE", "entry_ids": ["WMS-0001"],
            "feedback": "", "robs_answer": "Better answer here",
            "answer": "Old answer", "revision_hash": "",
            "question": "Q?", "skipped": False,
        }
        with patch("rfp_feedback.cmd_correct_offline", return_value=0) as mock_correct, \
             patch("rfp_feedback.check_forbidden_claims", return_value=[]), \
             patch("rfp_feedback.load_profile", return_value={}):
            result = execute_row(row, "wms", dry_run=False)
            mock_correct.assert_called_once_with("WMS-0001", "Better answer here", dry_run=False)
            assert result["parse_method"] == "direct"


# ---------------------------------------------------------------------------
# Stale detection
# ---------------------------------------------------------------------------

class TestStaleDetection:
    @patch("rfp_feedback.find_entry")
    @patch("rfp_feedback.load_entry")
    def test_stale_detected(self, mock_load, mock_find):
        mock_find.return_value = Path("/fake/entry.json")
        mock_load.return_value = {"answer": "new content changed"}
        from import_review_pack import _content_hash
        old_hash = _content_hash("original answer")
        assert check_stale(["WMS-0001"], old_hash) is True

    @patch("rfp_feedback.find_entry")
    @patch("rfp_feedback.load_entry")
    def test_fresh_passes(self, mock_load, mock_find):
        mock_find.return_value = Path("/fake/entry.json")
        answer = "same answer"
        mock_load.return_value = {"answer": answer}
        from import_review_pack import _content_hash
        h = _content_hash(answer)
        assert check_stale(["WMS-0001"], h) is False

    def test_empty_hash_not_stale(self):
        assert check_stale(["WMS-0001"], "") is False

    def test_no_entry_ids_not_stale(self):
        assert check_stale([], "abc123") is False


# ---------------------------------------------------------------------------
# Execution
# ---------------------------------------------------------------------------

class TestExecution:
    def _base_row(self, action="APPROVE", **kw):
        base = {
            "row": 2, "action": action, "entry_ids": ["WMS-0001"],
            "feedback": "", "robs_answer": "", "answer": "REST API",
            "revision_hash": "", "question": "Q?", "category": "technical",
            "topic": "integration", "skipped": False,
        }
        base.update(kw)
        return base

    @patch("rfp_feedback.find_entry_dir", return_value=(Path("/d/WMS-0001.json"), "drafts"))
    @patch("rfp_feedback.cmd_approve", return_value=0)
    @patch("rfp_feedback.check_forbidden_claims", return_value=[])
    @patch("rfp_feedback.load_profile", return_value={})
    def test_approve_moves_to_verified(self, _lp, _cfc, mock_approve, _fed):
        result = execute_row(self._base_row("APPROVE"), "wms", dry_run=False)
        assert result["status"] == "APPROVED"
        mock_approve.assert_called_once_with("WMS-0001")

    @patch("rfp_feedback.check_forbidden_claims", return_value=["On-premise only"])
    @patch("rfp_feedback.load_profile", return_value={})
    def test_quarantine_on_violation(self, _lp, _cfc):
        result = execute_row(self._base_row("APPROVE"), "wms", dry_run=False)
        assert result["status"] == "QUARANTINED"
        assert "Forbidden" in result["detail"]

    @patch("rfp_feedback.cmd_correct", return_value=0)
    @patch("rfp_feedback.check_forbidden_claims", return_value=[])
    @patch("rfp_feedback.load_profile", return_value={})
    def test_update_calls_correct(self, _lp, _cfc, mock_correct):
        row = self._base_row("UPDATE", feedback="Add more API details")
        result = execute_row(row, "wms", dry_run=False)
        assert result["status"] == "UPDATED"
        mock_correct.assert_called_once()

    @patch("rfp_feedback.cmd_correct_offline", return_value=0)
    @patch("rfp_feedback.check_forbidden_claims", return_value=[])
    @patch("rfp_feedback.load_profile", return_value={})
    def test_update_with_robs_answer(self, _lp, _cfc, mock_offline):
        row = self._base_row("UPDATE", robs_answer="Rob wrote this")
        result = execute_row(row, "wms", dry_run=False)
        assert result["status"] == "UPDATED"
        assert result["parse_method"] == "direct"
        mock_offline.assert_called_once_with("WMS-0001", "Rob wrote this", dry_run=False)

    @patch("rfp_feedback.find_entry_dir", return_value=(Path("/d/WMS-0001.json"), "drafts"))
    @patch("rfp_feedback.cmd_reject", return_value=0)
    @patch("rfp_feedback.check_forbidden_claims", return_value=[])
    @patch("rfp_feedback.load_profile", return_value={})
    def test_reject(self, _lp, _cfc, mock_reject, _fed):
        row = self._base_row("REJECT", feedback="Outdated info")
        result = execute_row(row, "wms", dry_run=False)
        assert result["status"] == "REJECTED"
        mock_reject.assert_called_once()

    @patch("rfp_feedback.save_entry")
    @patch("rfp_feedback.append_feedback_log")
    @patch("rfp_feedback.check_forbidden_claims", return_value=[])
    @patch("rfp_feedback.load_profile", return_value={})
    def test_new_creates_draft(self, _lp, _cfc, mock_log, mock_save, tmp_path):
        row = self._base_row("NEW", robs_answer="Brand new answer", entry_ids=[])
        with patch("import_review_pack.DRAFTS_DIR", tmp_path):
            result = execute_row(row, "wms", dry_run=False)
        assert result["status"] == "CREATED"
        mock_save.assert_called_once()

    def test_dry_run_no_changes(self):
        row = self._base_row("APPROVE")
        with patch("rfp_feedback.check_forbidden_claims", return_value=[]), \
             patch("rfp_feedback.load_profile", return_value={}):
            result = execute_row(row, "wms", dry_run=True)
        assert result["status"] == "WOULD_APPROVE"

    @patch("import_review_pack.check_stale", return_value=True)
    def test_conflict_on_stale(self, _cs):
        row = self._base_row("APPROVE", revision_hash="old_hash")
        result = execute_row(row, "wms", dry_run=False)
        assert result["status"] == "CONFLICT"


# ---------------------------------------------------------------------------
# General feedback
# ---------------------------------------------------------------------------

class TestGeneralFeedback:
    def test_saved_to_queue(self, tmp_path):
        xlsx = _create_pack_excel(
            tmp_path / "p.xlsx",
            general_fb=[{"feedback": "Improve security section", "scope": "product-wide"}],
        )
        items = parse_general_feedback(xlsx)
        assert len(items) == 1
        assert items[0]["feedback"] == "Improve security section"

    def test_empty_ignored(self, tmp_path):
        xlsx = _create_pack_excel(tmp_path / "p.xlsx", general_fb=[])
        items = parse_general_feedback(xlsx)
        assert len(items) == 0

    def test_placeholder_ignored(self, tmp_path):
        xlsx = _create_pack_excel(
            tmp_path / "p.xlsx",
            general_fb=[{"feedback": "(Write general feedback here)"}],
        )
        items = parse_general_feedback(xlsx)
        assert len(items) == 0


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

class TestReport:
    def test_shows_all_actions(self, capsys):
        summary = {
            "approved": 2, "updated": 1, "rejected": 1, "new": 1,
            "skipped": 3, "conflicts": 1, "quarantined": 1, "errors": 0,
            "total": 10, "actionable": 7,
            "update_rule_parsed": 1, "update_llm_parsed": 0,
            "general_feedback": 1,
        }
        results = [
            {"row": 2, "action": "APPROVE", "status": "APPROVED", "detail": "ok"},
            {"row": 3, "action": "REJECT", "status": "REJECTED", "detail": "bad"},
            {"row": 4, "action": "UPDATE", "status": "CONFLICT", "detail": "stale"},
        ]
        print_import_report("PACK001", "wms", summary, results, dry_run=False)
        output = capsys.readouterr().out
        assert "APPROVED" in output
        assert "REJECTED" in output
        assert "CONFLICTS" in output
        assert "QUARANTINED" in output
        assert "General Feedback" in output


# ---------------------------------------------------------------------------
# Tracking
# ---------------------------------------------------------------------------

class TestTracking:
    def test_processed_record_created(self, tmp_path):
        packs_file = tmp_path / "processed.jsonl"
        with patch("import_review_pack.PROCESSED_PACKS", packs_file):
            mark_processed("PACK001", "test.xlsx", {"approved": 1})
        assert packs_file.exists()
        record = json.loads(packs_file.read_text(encoding="utf-8").strip())
        assert record["pack_id"] == "PACK001"

    def test_reprocess_blocked_without_force(self, tmp_path):
        xlsx = _create_pack_excel(tmp_path / "pack.xlsx")
        packs_file = tmp_path / "processed.jsonl"
        packs_file.write_text(
            json.dumps({"pack_id": "WMS_20260312_PACK001"}) + "\n", encoding="utf-8"
        )
        with patch("import_review_pack.PROCESSED_PACKS", packs_file):
            summary = import_pack(xlsx, dry_run=True, force=False)
        # Should short-circuit with empty summary
        assert summary["total"] == 0
