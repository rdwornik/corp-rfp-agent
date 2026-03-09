"""Tests for AnswerWriter."""

import pytest
from pathlib import Path
from openpyxl import load_workbook

from corp_rfp_agent.agents.excel.answer_writer import AnswerWriter
from corp_rfp_agent.agents.excel.models import GreenCell

FIXTURES_DIR = Path(__file__).resolve().parents[1] / "fixtures"


@pytest.fixture
def golden_wb():
    return load_workbook(str(FIXTURES_DIR / "excel_golden.xlsx"), keep_links=True)


def test_write_answer_sets_value(golden_wb, tmp_path):
    """write_answer sets cell value."""
    writer = AnswerWriter(golden_wb)
    cell = GreenCell(
        sheet_name="Requirements", row=2, answer_col_idx=5,
        question_col_idx=3, question_text="Q?"
    )
    writer.write_answer(cell, "Test answer")
    writer.save(tmp_path / "out.xlsx")

    wb2 = load_workbook(str(tmp_path / "out.xlsx"))
    assert wb2["Requirements"].cell(row=2, column=5).value == "Test answer"


def test_formula_preserved_after_write(golden_wb, tmp_path):
    """Formula in F8 is preserved after writing to E8."""
    writer = AnswerWriter(golden_wb)
    cell = GreenCell(
        sheet_name="Requirements", row=8, answer_col_idx=5,
        question_col_idx=3, question_text="Q?"
    )
    writer.write_answer(cell, "Scaling answer")
    writer.save(tmp_path / "out.xlsx")

    wb2 = load_workbook(str(tmp_path / "out.xlsx"))
    assert wb2["Requirements"]["F8"].value == "=LEN(E8)"


def test_answered_cells_unchanged(golden_wb, tmp_path):
    """Already-answered cells are not modified."""
    writer = AnswerWriter(golden_wb)
    # Write to row 2 only
    cell = GreenCell(
        sheet_name="Requirements", row=2, answer_col_idx=5,
        question_col_idx=3, question_text="Q?"
    )
    writer.write_answer(cell, "New answer")
    writer.save(tmp_path / "out.xlsx")

    wb2 = load_workbook(str(tmp_path / "out.xlsx"))
    ws = wb2["Requirements"]
    # Row 4 answer should be unchanged
    assert ws.cell(row=4, column=5).value == "Blue Yonder supports cloud-native SaaS deployment on Microsoft Azure."
    # Row 6 answer should be unchanged
    assert ws.cell(row=6, column=5).value == "Yes, the platform is fully GDPR compliant with data residency options."


def test_save_produces_valid_workbook(golden_wb, tmp_path):
    """Saved workbook opens without error."""
    writer = AnswerWriter(golden_wb)
    writer.save(tmp_path / "out.xlsx")

    wb2 = load_workbook(str(tmp_path / "out.xlsx"))
    assert "Requirements" in wb2.sheetnames
    assert wb2["Requirements"].max_row >= 10


def test_merged_cells_survive_write(golden_wb, tmp_path):
    """Merged cells survive write + save cycle."""
    writer = AnswerWriter(golden_wb)
    cell = GreenCell(
        sheet_name="Requirements", row=7, answer_col_idx=5,
        question_col_idx=3, question_text="Q?"
    )
    writer.write_answer(cell, "Performance answer")
    writer.save(tmp_path / "out.xlsx")

    wb2 = load_workbook(str(tmp_path / "out.xlsx"))
    ws = wb2["Requirements"]
    # Merged range should still exist
    merged = [str(m) for m in ws.merged_cells.ranges]
    assert any("C7" in m for m in merged)
