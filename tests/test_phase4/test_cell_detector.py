"""Tests for CellDetector -- validates against golden fixture.

These tests verify that v2 CellDetector finds the exact same cells
as the v1 agent from src/rfp_excel_agent.py.
"""

import pytest
from pathlib import Path

from corp_rfp_agent.agents.excel.cell_detector import (
    CellDetector,
    is_green_cell,
    find_header_row,
    detect_question_column,
    detect_answer_column,
)

FIXTURES_DIR = Path(__file__).resolve().parents[1] / "fixtures"


@pytest.fixture
def golden_path():
    return FIXTURES_DIR / "excel_golden.xlsx"


@pytest.fixture
def detector(golden_path):
    return CellDetector(golden_path)


def test_green_cells_detected(detector):
    """Finds exactly 6 green cells at expected rows."""
    cells = detector.scan_green_cells()
    found_rows = sorted([c.row for c in cells])
    assert found_rows == [2, 3, 5, 7, 8, 9]


def test_answered_cells_not_detected(detector):
    """Rows 4 and 6 (already answered) are NOT detected."""
    cells = detector.scan_green_cells()
    found_rows = [c.row for c in cells]
    assert 4 not in found_rows
    assert 6 not in found_rows


def test_category_header_not_detected(detector):
    """Row 10 (category header, not green) is NOT detected."""
    cells = detector.scan_green_cells()
    found_rows = [c.row for c in cells]
    assert 10 not in found_rows


def test_question_text_extracted(detector):
    """Question text is correctly extracted from question column."""
    cells = detector.scan_green_cells()
    cell_map = {c.row: c for c in cells}

    assert "REST API" in cell_map[2].question_text
    assert "SSO" in cell_map[3].question_text
    assert "encryption" in cell_map[5].question_text


def test_merged_cell_handled(detector):
    """Row 7 merged cell question text is extracted."""
    cells = detector.scan_green_cells()
    row7 = [c for c in cells if c.row == 7]
    assert len(row7) == 1
    assert "concurrent users" in row7[0].question_text


def test_special_characters_preserved(detector):
    """Row 9 special characters survive detection."""
    cells = detector.scan_green_cells()
    row9 = [c for c in cells if c.row == 9]
    assert len(row9) == 1
    assert '"quotes"' in row9[0].question_text
    assert "&" in row9[0].question_text


def test_header_row_detection(golden_path):
    """Header row is detected at row 1."""
    from openpyxl import load_workbook
    wb = load_workbook(str(golden_path))
    ws = wb["Requirements"]
    assert find_header_row(ws) == 1


def test_question_column_detection(golden_path):
    """Question column is column C (3) via 'Requirement' header."""
    from openpyxl import load_workbook
    wb = load_workbook(str(golden_path))
    ws = wb["Requirements"]
    header_row = find_header_row(ws)
    col_idx, col_name = detect_question_column(ws, header_row)
    assert col_idx == 3
    assert "Requirement" in col_name


def test_answer_column_detection(golden_path):
    """Answer column is column E (5) via 'Vendor Response' header."""
    from openpyxl import load_workbook
    wb = load_workbook(str(golden_path))
    ws = wb["Requirements"]
    header_row = find_header_row(ws)
    q_col, _ = detect_question_column(ws, header_row)
    a_col, a_name = detect_answer_column(ws, header_row, q_col)
    assert a_col == 5
    assert "response" in a_name.lower() or "vendor" in a_name.lower()


def test_is_green_cell_exact_match(golden_path):
    """is_green_cell matches exact FF00FF00 color."""
    from openpyxl import load_workbook
    wb = load_workbook(str(golden_path))
    ws = wb["Requirements"]

    # E2 is green
    assert is_green_cell(ws["E2"]) is True
    # E4 is not green (has answer, no green fill)
    assert is_green_cell(ws["E4"]) is False
    # A1 is not green (header fill)
    assert is_green_cell(ws["A1"]) is False
