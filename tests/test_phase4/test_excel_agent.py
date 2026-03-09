"""Tests for ExcelAgent v2 with mock LLM and KB."""

import pytest
from pathlib import Path
from dataclasses import dataclass, field
from openpyxl import load_workbook

from corp_rfp_agent.agents.excel.agent import ExcelAgent
from corp_rfp_agent.agents.excel.models import ExcelAgentResult
from corp_rfp_agent.anonymize.service import Anonymizer
from corp_rfp_agent.core.types import LLMResponse, KBMatch

FIXTURES_DIR = Path(__file__).resolve().parents[1] / "fixtures"


class MockLLMClient:
    """Mock LLM that returns a fixed answer."""
    def __init__(self, answer="Blue Yonder supports this capability."):
        self._answer = answer
        self.calls = []

    def generate(self, prompt, *, model=None, system_prompt=None, temperature=0.3, max_tokens=2000):
        self.calls.append(prompt)
        return LLMResponse(text=self._answer, model=model or "mock", provider="mock")

    def generate_json(self, prompt, *, model=None, system_prompt=None):
        return {}


class MockKBClient:
    """Mock KB that returns fixed matches."""
    def __init__(self, matches=None):
        self._matches = matches or []

    def query(self, question, *, family=None, category=None, top_k=5, threshold=0.75):
        return self._matches

    def count(self, family=None):
        return len(self._matches)

    def families(self):
        return {}


def test_process_dry_run():
    """Dry run returns green cells without generating answers."""
    agent = ExcelAgent(
        llm_client=MockLLMClient(),
        kb_client=MockKBClient(),
    )
    result = agent.process(
        FIXTURES_DIR / "excel_golden.xlsx",
        dry_run=True,
    )
    assert result.total_green_cells == 6
    assert result.answered == 0


def test_process_writes_answers(tmp_path):
    """Process with mock LLM writes answers to correct cells."""
    agent = ExcelAgent(
        llm_client=MockLLMClient(answer="Mock answer for testing."),
        kb_client=MockKBClient(),
    )
    output = tmp_path / "output.xlsx"
    result = agent.process(
        FIXTURES_DIR / "excel_golden.xlsx",
        output_path=output,
    )
    assert result.answered == 6
    assert result.errors == 0
    assert output.exists()

    # Verify answers written
    wb = load_workbook(str(output))
    ws = wb["Requirements"]
    assert ws.cell(row=2, column=5).value == "Mock answer for testing."
    assert ws.cell(row=3, column=5).value == "Mock answer for testing."


def test_process_preserves_answered_cells(tmp_path):
    """Existing answers (rows 4, 6) are preserved."""
    agent = ExcelAgent(
        llm_client=MockLLMClient(answer="New answer"),
        kb_client=MockKBClient(),
    )
    output = tmp_path / "output.xlsx"
    agent.process(FIXTURES_DIR / "excel_golden.xlsx", output_path=output)

    wb = load_workbook(str(output))
    ws = wb["Requirements"]
    assert "cloud-native" in ws.cell(row=4, column=5).value
    assert "GDPR" in ws.cell(row=6, column=5).value


def test_process_preserves_formula(tmp_path):
    """Formula in F8 is preserved after processing."""
    agent = ExcelAgent(
        llm_client=MockLLMClient(answer="Answer"),
        kb_client=MockKBClient(),
    )
    output = tmp_path / "output.xlsx"
    agent.process(FIXTURES_DIR / "excel_golden.xlsx", output_path=output)

    wb = load_workbook(str(output))
    assert wb["Requirements"]["F8"].value == "=LEN(E8)"


def test_process_applies_anonymization(tmp_path):
    """Anonymizer is called before LLM."""
    llm = MockLLMClient(answer="Answer for [Customer]")
    anonymizer = Anonymizer(client_name="AcmeCorp")
    agent = ExcelAgent(
        llm_client=llm,
        kb_client=MockKBClient(),
        anonymizer=anonymizer,
    )
    output = tmp_path / "output.xlsx"
    agent.process(FIXTURES_DIR / "excel_golden.xlsx", output_path=output)

    # LLM was called
    assert len(llm.calls) == 6
    # De-anonymized answer should have AcmeCorp restored
    # (but original questions don't contain "AcmeCorp" so anonymizer is a no-op here)


def test_process_applies_overrides(tmp_path):
    """Overrides are applied to final answer."""
    import yaml
    overrides_path = tmp_path / "overrides.yaml"
    data = {"overrides": [
        {"id": "OVR-TEST", "find": "Mock answer", "replace": "Overridden answer"},
    ]}
    overrides_path.write_text(yaml.dump(data), encoding="utf-8")

    from corp_rfp_agent.overrides.store import YAMLOverrideStore
    store = YAMLOverrideStore(yaml_path=overrides_path)

    agent = ExcelAgent(
        llm_client=MockLLMClient(answer="Mock answer for testing."),
        kb_client=MockKBClient(),
        override_store=store,
    )
    output = tmp_path / "output.xlsx"
    result = agent.process(FIXTURES_DIR / "excel_golden.xlsx", output_path=output)

    # All answers should have the override applied
    wb = load_workbook(str(output))
    ws = wb["Requirements"]
    assert "Overridden answer" in ws.cell(row=2, column=5).value


def test_result_tracks_counts(tmp_path):
    """ExcelAgentResult tracks correct counts."""
    agent = ExcelAgent(
        llm_client=MockLLMClient(answer="A"),
        kb_client=MockKBClient(),
    )
    output = tmp_path / "output.xlsx"
    result = agent.process(FIXTURES_DIR / "excel_golden.xlsx", output_path=output)

    assert result.total_green_cells == 6
    assert result.answered == 6
    assert result.skipped == 0
    assert result.errors == 0
    assert len(result.results) == 6
